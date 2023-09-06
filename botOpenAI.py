import telebot
import sqlite3
from telebot import types
from telegram import ParseMode
import datetime
import urllib.request
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram import ReplyKeyboardMarkup, KeyboardButton
from telebot.types import ReplyKeyboardMarkup, KeyboardButton
import os
import openai
from telebot import TeleBot
import urllib.request
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, Update
from telegram import Update
from telegram.ext import Updater, CommandHandler, CallbackContext, CallbackQueryHandler, ConversationHandler
from telegram.ext import CommandHandler
import time
import json
import openpyxl

from telegram import InputFile

####5428270852:AAEbBDt8RiYgiizDEC7o5oTz4vl-x7Ls5ng    тестовий

#reply_markup=telebot.types.ReplyKeyboardRemove()
#5646599316:AAFVGWqEAgPmlvpUByhFwmbDjB-1UFY7LWY      основний бот
CHAT_ID = 628446966
TELEGRAM_API_KEY = '5646599316:AAFVGWqEAgPmlvpUByhFwmbDjB-1UFY7LWY'
openai.api_key = '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'
bot = telebot.TeleBot(TELEGRAM_API_KEY)
def get_user_data(user_id):
    connect = sqlite3.connect('users.db')
    cursor = connect.cursor()
    cursor.execute(f"SELECT grypa, email, first_last FROM login_id WHERE id = {user_id}")
    data = cursor.fetchone()
    connect.close()
    return data
@bot.message_handler(commands=['start'])
def start(message: types.Message):
    connect = sqlite3.connect('users.db')
    cursor = connect.cursor()

    cursor.execute("""CREATE TABLE IF NOT EXISTS login_id(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT,
            email TEXT NOT NULL,
            grypa TEXT NOT NULL,
            first_last TEXT NOT NULL,
            roli TEXT )""")

    connect.commit()
    people_id = message.chat.id
    cursor.execute(f"SELECT id FROM login_id WHERE id = {people_id} ")
    data = cursor.fetchone()
    if data is None:

        # Запит email
        bot.send_message(message.chat.id, "Будь ласка, введіть свою email адресу:")
        bot.register_next_step_handler(message, get_email)

    else:

        bot.send_message(
            chat_id=message.chat.id,
            text="Ти вже зареєстрований!",
            reply_to_message_id=message.message_id
        )
        time.sleep(2)
        message_handler_start(message)
def get_email(message: types.Message):
    email = message.text
    email = email.lower()
    connect = sqlite3.connect('users.db')
    cursor = connect.cursor()
    cursor.execute("SELECT email FROM login_id")
    rows = cursor.fetchall()
    emails = [row[0] for row in rows]  # Створення списку зі значень email

    if email in emails:
        bot.send_message(message.chat.id, "Користувач за такою електронною адресою вже зареєстрований введіть будь ласка свою адресу")
        bot.register_next_step_handler(message, get_email)

    elif email.startswith('/'):
        bot.send_message(message.chat.id, "Ви ввели команду а не пошту будь ласка введіть свою фізматівську пошту: ")
        bot.register_next_step_handler(message, get_email)

    else:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Email_Base WHERE Email_Address=?", (email,))
        row = cursor.fetchone()
        cursor.execute("SELECT ПІП FROM Email_Base WHERE Email_Address=?", (email,))
        pib = cursor.fetchone()

        if row:
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            button2 = types.KeyboardButton("Так")
            button1 = types.KeyboardButton("Ні")
            keyboard.add(button2, button1)
            bot.send_message(message.chat.id, f"{pib[0]}\nБудь ласка скажіть це ваше прізвище?", reply_markup=keyboard)
            bot.register_next_step_handler(message, get_first_last, email,pib)


        else:
            # Email does not exist, send error message
            bot.send_message(message.chat.id,"Вашої пошти не знайдено в базі даних фізмату. Будь ласка, введіть свою пошту ще раз правильно.")
            # Return to the get_email function to wait for the next input from the user
            bot.register_next_step_handler(message, get_email)
        cursor.close()
        conn.close()
def get_first_last(message,email,pib):
    text = message.text
    if text == "Так":
        keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
        button_student = types.KeyboardButton("Студент")
        button_starosta = types.KeyboardButton("Староста")
        button_vikladach = types.KeyboardButton("Викладач")
        keyboard.add(button_student, button_starosta, button_vikladach)
        bot.send_message(message.chat.id, "Виберіть вашу роль:", reply_markup=keyboard)
        bot.register_next_step_handler(message, get_role, email,pib)

    elif text =="Ні":
        bot.send_message(message.chat.id, "Введіть тоді своє ПІБ:")
        bot.register_next_step_handler(message, get_first_last_2, email, pib)

    else:
        bot.send_message(message.chat.id,"Такого варіанту немає")
        bot.register_next_step_handler(message, get_first_last, email, pib)

def get_first_last_2(message,email,pib):
    text = message.text
    pib = text
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button_student = types.KeyboardButton("Студент")
    button_starosta = types.KeyboardButton("Староста")
    button_vikladach = types.KeyboardButton("Викладач")
    keyboard.add(button_student, button_starosta, button_vikladach)
    bot.send_message(message.chat.id, "Виберіть вашу роль:", reply_markup=keyboard)
    bot.register_next_step_handler(message, get_role, email, pib)
def get_role(message, email,pib):
    role = message.text

    if role.startswith('/'):
        bot.send_message(message.chat.id,"Ви ввели команду, а не роль")
        bot.register_next_step_handler(message, get_role, email,pib)

    elif role == 'Студент':
        connect = sqlite3.connect('users.db')
        cursor = connect.cursor()
        cursor.execute("SELECT Групи FROM Групи")
        rows = cursor.fetchall()
        gryps = [row[0] for row in rows]
        columns = 3
        gryps_per_column = (len(gryps) + columns - 1) // columns
        gryps_divided = [gryps[i:i + gryps_per_column] for i in range(0, len(gryps), gryps_per_column)]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

        # Додавання груп до розмітки по 3 в кожному рядку
        for gryp_column in gryps_divided:
            markup.add(*gryp_column)

        connect.close()

        bot.send_message(message.chat.id,
                         "Будь ласка, будьте уважні при виборі своєї групи. Оберіть дійсну групу, оскільки редагування групи не буде можливим. Якщо помилилися з вибором групи напишіть в /support",
                         reply_markup=markup)
        bot.register_next_step_handler(message, get_group_stud, email,role,pib, gryps)

    elif role == 'Викладач':
        # Запит паролю для ролі викладача
        bot.send_message(message.chat.id, "Будь ласка, введіть пароль для викладача:")
        bot.register_next_step_handler(message, get_password,  role, email,pib)

    elif role == 'Староста':
        # Запит паролю для ролі старости
        bot.send_message(message.chat.id, "Будь ласка, введіть пароль для старости:")
        bot.register_next_step_handler(message, get_password, role, email,pib)

    else:
        # Надсилання повідомлення про неправильний вибір ролі
        bot.send_message(message.chat.id, "Виберіть роль з наданих кнопок.")
        bot.register_next_step_handler(message, get_role, email,pib)
def get_password(message, role, email,pib):
    password = message.text

    if role == 'Викладач' and password == '0x966a605b5D315871e05acb788193aE4f576435B2':
        board = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Математика📏')
        item2 = types.KeyboardButton('Інформатика🧑‍💻')
        item3 = types.KeyboardButton('Фізика👨‍🔬')
        board.add(item1, item2, item3)
        bot.send_message(message.chat.id, "Виберіть із якої ви кафедри:", reply_markup=board)
        bot.register_next_step_handler(message, kafedra, email, role,pib)

    elif role == 'Староста' and password == 'bc1q9p6rys9x242akrwgyw35lxxe35pd7tagtk3zhs':
        connect = sqlite3.connect('users.db')
        cursor = connect.cursor()
        cursor.execute("SELECT Групи FROM Групи")
        rows = cursor.fetchall()
        gryps = [row[0] for row in rows]
        columns = 3
        gryps_per_column = (len(gryps) + columns - 1) // columns
        gryps_divided = [gryps[i:i + gryps_per_column] for i in range(0, len(gryps), gryps_per_column)]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)

        # Додавання груп до розмітки по 3 в кожному рядку
        for gryp_column in gryps_divided:
            markup.add(*gryp_column)

        connect.close()

        bot.send_message(message.chat.id,
                         "Будь ласка, будьте уважні при виборі своєї групи. Оберіть дійсну групу, оскільки редагування групи не буде можливим. Якщо помилилися з вибором групи напишіть в /support",
                         reply_markup=markup)
        bot.register_next_step_handler(message, get_group_stud, email, role, pib, gryps)




    else:
        # Надсилання повідомлення про неправильний пароль
        bot.send_message(message.chat.id, "Неправильний пароль. Спробуйте ще раз.")
        bot.register_next_step_handler(message, get_password, role, email,pib)
def kafedra(message, email, role,pib):
    group = message.text
    if group not in ['Математика📏', 'Інформатика🧑‍💻', 'Фізика👨‍🔬']:
        bot.send_message(message.chat.id, "Ви що з хімбіо?🤨")
        bot.register_next_step_handler(message, kafedra, message, email, role,pib)

    elif group.startswith('/'):
        bot.send_message(message.chat.id, "Будь ласка будьте уважніші ви ввели команду, а не назву кафедри, будь ласка введіть свою кафедру😡 ")
        bot.register_next_step_handler(message, kafedra, email, role,pib)
    else:

        if group == 'Математика📏':
            group = "Математик"
        elif group == 'Інформатика🧑‍💻':
            group = 'Інформатик'
        elif group == 'Фізика👨‍🔬':
            group = 'Фізик'
        get(message, email, role, group, pib)
def get_group_stud(message,email,role,pib,gryps):
    group = message.text
    if group not in gryps:
        bot.send_message(message.chat.id, "Ви ввели не правильну групу виберіть ще раз свою групу:")
        bot.register_next_step_handler(message, get_group_stud, email, role,pib,gryps)
    elif group.startswith('/'):
        bot.send_message(message.chat.id, "Будь ласка будьте уважніші ви ввели команду а не назву групи, будь ласка введіть свою групу😡 ")
        bot.register_next_step_handler(message, get_group_stud, email,role,pib,gryps)
    else:
        group = message.text.upper().replace('-', '_')
        get(message, email, role, group,pib)
def get(message, email, role, group,pib):
    role = role.lower()
    email = email.lower()

    connect = sqlite3.connect('users.db')
    cursor = connect.cursor()
    pib = pib[0]
    user_id = message.chat.id
    user_name = message.chat.username
    
    # Вставка даних в базу даних
    cursor.execute("INSERT INTO login_id (id, username, email, grypa, first_last, roli) VALUES (?, ?, ?, ?, ?, ?);",
                   (user_id, user_name, email, group, pib, role))
    connect.commit()
    bot.send_message(message.chat.id, "Успішна реєстрація")
    create_rozklad_table(message)
def create_rozklad_table(message):
    # Встановлення з'єднання з базою даних
    connect = sqlite3.connect('users.db')
    cursor = connect.cursor()
    connect = sqlite3.connect('users.db')
    cursor = connect.cursor()
    cursor.execute("SELECT Групи FROM Групи")
    rows = cursor.fetchall()
    groups = [row[0] for row in rows]

    # Створення таблиць для кожної групи
    for group in groups:
        table_name = f"rosklad_{group.replace('-', '_')}"
        # Перевірка наявності таблиці
        cursor.execute(f"SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='{table_name}'")
        result = cursor.fetchone()
        if result[0] == 0:
            # Створення таблиці, якщо вона не існує
            cursor.execute(
                f"CREATE TABLE {table_name} (Перша TEXT, друга TEXT, третя TEXT, четверта TEXT, пята TEXT)")
            # Додавання значення "пара ще не вказана" у всі стовпці таблиці
            cursor.execute(
                f"INSERT INTO {table_name} (Перша, друга, третя, четверта, пята) VALUES ('пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана')")
            cursor.execute(
                f"INSERT INTO {table_name} (Перша, друга, третя, четверта, пята) VALUES ('пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана')")
            cursor.execute(
                f"INSERT INTO {table_name} (Перша, друга, третя, четверта, пята) VALUES ('пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана')")
            cursor.execute(
                f"INSERT INTO {table_name} (Перша, друга, третя, четверта, пята) VALUES ('пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана')")
            cursor.execute(
                f"INSERT INTO {table_name} (Перша, друга, третя, четверта, пята) VALUES ('пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана', 'пара ще не вказана')")

    connect.commit()  # Застосовуємо зміни до бази даних
    cursor.close()
    connect.close()
    message_handler_start(message)
###############################################################################################################################################################################
@bot.message_handler(commands=['shurik'])
def shurik(message):
    bot.send_message(message.chat.id, "Помідор ваше пєрсік, мамою твоєю клянусь".format(message.from_user))
@bot.message_handler(commands=['legion'])
def legion(message):
    bot.send_message(message.chat.id, "пшш пшш пшш Олег пукнув\nце Олег @phantomkahueta ".format(message.from_user))
#####################################################################################################################

######################################################################################################################

@bot.message_handler(commands=['delete'])
def delete(message):
    # запитуємо у користувача айді отримувача повідомлення
    chat_id = message.chat.id
    if chat_id != CHAT_ID:
        bot.send_message(chat_id=chat_id, text='Ви не маєте доступу до цієї команди.')
        return
    else:
        bot.send_message(message.chat.id, "Введіть айді користувача: ")
        bot.register_next_step_handler(message, handle_user_id)
def handle_user_id(message):
    # Отримуємо введене айді користувача
    user_id = message.text

    # Викликаємо функцію для видалення користувача з бази даних
    delete_user(user_id)

    # Відправляємо підтвердження видалення користувача
    bot.send_message(message.chat.id, f"Користувач з айді {user_id} видалений.")
def delete_user(user_id):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    # Видаляємо користувача з бази даних
    cursor.execute("DELETE FROM login_id WHERE id=?", (user_id,))
    conn.commit()
@bot.message_handler(commands=['support'])
def message_handler_support(message):
    # Відправлення повідомлення від бота з кнопкою "🔙Назад"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_back = types.KeyboardButton('🔙Назад')
    markup.add(item_back)
    bot.send_message(chat_id=message.chat.id,
                     text='<b>Доброго дня, будь ласка опишіть детально з якими труднощами ви зіткнулися?</b>',
                     parse_mode=ParseMode.HTML, reply_markup=markup)
    # Реєстрація наступного кроку з обробником повідомлення користувача
    bot.register_next_step_handler(message, support_reply_handler)
def support_reply_handler(message):
    if message.text == '🔙Назад':
        # Виклик команди "/start" при натисканні кнопки "🔙Назад"
        message_handler_start(message)
    else:
        # Відправка повідомлення користувачем до підтримки
        bot.send_message(chat_id=628446966,
                         text=f'Користувач звернувся за допомогою:\nАйді: {message.chat.id}\nНік: @{message.chat.username}\nТекст: {message.text}')
        bot.send_message(chat_id=message.chat.id,
                         text='Дякуємо за ваше повідомлення! Наша команда підтримки зв\'яжеться з вами найближчим часом.')
        message_handler_start(message)
@bot.message_handler(commands=['idea'])
def message_handler_idea(message):
    # Відправлення повідомлення від бота з кнопкою "🔙Назад"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_back = types.KeyboardButton('🔙Назад')
    markup.add(item_back)
    bot.send_message(chat_id=message.chat.id,
                     text='<b>Яка у вас є ідея для мене? Напишіть її, і ми розглянемо її реалізацію в майбутньому.</b>',
                     parse_mode=ParseMode.HTML, reply_markup=markup)
    # Реєстрація наступного кроку з обробником повідомлення користувача
    bot.register_next_step_handler(message, idea_reply_handler)
def idea_reply_handler(message):
    if message.text == '🔙Назад':
        # Виклик команди "/start" при натисканні кнопки "🔙Назад"
        message_handler_start(message)
    else:
        # Відправка ідеї користувачем
        bot.send_message(chat_id=628446966,
                         text=f'Користувач запропонував ідею для бота:\nАйді: {message.chat.id}\nНік: @{message.chat.username}\nТекст: {message.text}')
        bot.send_message(chat_id=message.chat.id,
                         text='Дякуємо за вашу ідею! Ми розглянемо її в майбутньому.')
@bot.message_handler(commands=['userhelp'])
def send_help(message):
    # запитуємо у користувача айді отримувача повідомлення
    uhat_id = message.chat.id
    if uhat_id != CHAT_ID:
        bot.send_message(chat_id=uhat_id, text='Ви не маєте доступу до цієї команди.')
        return
    else:
        bot.send_message(message.chat.id, "Введіть айді користувача: ")
        bot.register_next_step_handler(message, get_recipient_id)
def get_recipient_id(message):
    # зберігаємо айді отримувача повідомлення та запитуємо текст повідомлення
    recipient_id = message.text
    bot.send_message(message.chat.id, "Введіть текст повідомлення:")
    bot.register_next_step_handler(message, send_message, recipient_id)
def send_message(message, recipient_id):
    # відправляємо повідомлення з введеним текстом до користувача з введеним айді
    try:
        bot.send_message(recipient_id, message.text)
        bot.send_message(message.chat.id, "Повідомлення надіслано!")
    except Exception as e:
        bot.send_message(message.chat.id, f"Помилка: {str(e)}")
@bot.message_handler(content_types=['voice', 'video_note'])
def handle_message(message):
    channel_id = '-1001955388901'
    if message.voice:
        # Отримано голосове повідомлення
        bot.forward_message(chat_id=channel_id, from_chat_id=message.chat.id, message_id=message.message_id)
    elif message.video_note:
        # Отримано кружечок повідомлення
        bot.forward_message(chat_id=channel_id, from_chat_id=message.chat.id, message_id=message.message_id)
@bot.message_handler(commands=['news'])
def send_news(message):
    # Запит повідомлення, яке потрібно розіслати
    bot.send_message(chat_id=message.chat.id, text='Введіть повідомлення для розсилки:')
    bot.register_next_step_handler(message, news_handler)
def news_handler(message):
    # Отримання повідомлення від користувача та розсилка його всім користувачам бота
    news = message.text
    users = get_all_users()  # Отримання всіх користувачів бота
    blocked_users = []
    for user in users:
        try:
            bot.send_message(chat_id=user, text=news)
        except telebot.apihelper.ApiTelegramException as e:
            if e.result.status_code == 403:
                blocked_users.append(user)
    # Перевірка, чи є заблоковані користувачі
    if blocked_users:
        # Формування повідомлення про заблокованих користувачів
        blocked_users_text = '\n'.join([str(user) for user in blocked_users])
        error_message = f"Користувачі {blocked_users_text} заблокували бота."
        # Надсилання повідомлення про заблокованих користувачів адміністратору бота
        bot.send_message(chat_id=CHAT_ID, text=error_message)
def get_all_users():
    # Підключення до бази даних
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    # Отримання списку всіх користувачів з таблиці login_id
    cursor.execute("SELECT id FROM login_id")
    users = cursor.fetchall()
    # Закриття підключення до бази даних
    cursor.close()
    conn.close()
    # Повернення списку всіх користувачів
    return [user[0] for user in users]
def menu_starostam(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton('Редагувати розклад')
    item2 = types.KeyboardButton('Робота з журналом')
    homework = types.KeyboardButton('Додати домашку')
    item4 = types.KeyboardButton('Оголошення для групи')
    back = types.KeyboardButton('🔙Назад')

    markup.add(back)
    markup.add(item1)
    markup.add(item2)
    markup.add(homework)
    markup.add(item4)
    bot.send_message(message.chat.id, "Це меню призначене спеціально для старост, ви можете ознайомитися зі списком функцій та можливостей у ньому".format(message.from_user), reply_markup=markup)
def menu_vikladacham(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item2 = types.KeyboardButton('Робота з журналами')
    homework = types.KeyboardButton('Додати домашнє')
    item4 = types.KeyboardButton('Оголошення для групи')
    back = types.KeyboardButton('🔙Назад')
    markup.add(back)
    markup.add(item2)
    markup.add(homework)
    markup.add(item4)
    bot.send_message(message.chat.id,"Ви попали в меню викладачам оберіть функцію з якою хочете працювати".format(message.from_user), reply_markup=markup)
    bot.register_next_step_handler(message,menu_vikladacham_2)
def menu_vikladacham_2(message):
    text = message.text
    if text == 'Робота з журналами':
        menu_vikladacham_3_1(message)

    elif text == 'Додати домашнє':
        bot.send_message(message.chat.id, "Ця функція поки що в розробці")
        bot.register_next_step_handler(message, menu_vikladacham_2)

    elif text == 'Оголошення для групи':
        connect = sqlite3.connect('users.db')
        cursor = connect.cursor()
        cursor.execute("SELECT Групи FROM Групи")
        rows = cursor.fetchall()
        gryps = [row[0] for row in rows]
        columns = 3
        gryps_per_column = (len(gryps) + columns - 1) // columns
        gryps_divided = [gryps[i:i + gryps_per_column] for i in range(0, len(gryps), gryps_per_column)]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        # Додавання груп до розмітки по 3 в кожному рядку
        for gryp_column in gryps_divided:
            markup.add(*gryp_column)

        connect.close()

        bot.send_message(message.chat.id, "Оберіть групу, якій хочете розіслати оголошення", reply_markup=markup)
        bot.register_next_step_handler(message, ogolosh_grypam,gryps)

    elif text == '🔙Назад':
        message_handler_start(message)
    else:
        bot.send_message(message.chat.id, "Такого варінту відповіді немає(")
        bot.register_next_step_handler(message, menu_vikladacham_2)

def ogolosh_grypam(message,gryps):
    text = message.text
    if text == '🔙Назад':
        menu_vikladacham(message)
    elif text in gryps:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        back = types.KeyboardButton('🔙Назад')
        markup.add(back)
        bot.send_message(message.chat.id, f"Надішліть повідомлення, яке хочете розіслати групі {text}",reply_markup=markup)
        bot.register_next_step_handler(message, ogolosh_grypam_1,text)
    else:
        bot.send_message(message.chat.id, f"Такого варіанту немає оберіть ще раз")
        bot.register_next_step_handler(message, ogolosh_grypam, gryps)

def ogolosh_grypam_1(message,text):
    news = message.text

    if news == '🔙Назад':
        message_handler_start(message)

    else:
        news = message.text
        text = text.replace("-","_")
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT id FROM login_id WHERE grypa = '{text}'")
        users = cursor.fetchall()
        users = [item for tpl in users for item in tpl]
        blocked_users = []
        for user in users:
            try:
                bot.forward_message(chat_id=user, from_chat_id=message.chat.id, message_id=message.message_id)
            except telebot.apihelper.ApiTelegramException as e:
                if e.result.status_code == 403:
                    blocked_users.append(user)
        # Перевірка, чи є заблоковані користувачі
        if blocked_users:
            # Формування повідомлення про заблокованих користувачів
            blocked_users_text = '\n'.join([str(user) for user in blocked_users])
            error_message = f"Користувачі {blocked_users_text} заблокували бота."
            # Надсилання повідомлення про заблокованих користувачів адміністратору бота
            bot.send_message(chat_id=CHAT_ID, text=error_message)
        menu_vikladacham(message)


def menu_vikladacham_3_1(message):
    user_id = message.from_user.id

    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f'SELECT DISTINCT група FROM "{user_id}_tea"')
    res = cursor.fetchall()
    gryp = [item for tpl in res for item in tpl]
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('🔙Назад'))
    for i in gryp:
        markup.add(i)
    conn.close()
    bot.send_message(message.chat.id, "Оберіть групу з якою хочете працювати", reply_markup=markup)
    bot.register_next_step_handler(message, menu_vikladacham_3, gryp, user_id)
def menu_vikladacham_3(message,gryp,user_id):
    text = message.text
    if text == '🔙Назад':
        menu_vikladacham_2(message)
    elif text in gryp:
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT предмет FROM '{user_id}_tea' WHERE група = ?",(text,))
        res = cursor.fetchall()
        res = [item for tpl in res for item in tpl]
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        for i in res:
            markup.add(i)
        conn.close()
        bot.send_message(message.chat.id, f"Оберіть предмет, який хочете редагувати у групі {text}",reply_markup=markup)
        grypa = text
        bot.register_next_step_handler(message, menu_vikladacham_4, grypa, user_id)
    else:
        bot.send_message(message.chat.id, "Такого варінту немає оберіть групу ще раз")
        bot.register_next_step_handler(message, menu_vikladacham_3, gryp, user_id)
def menu_vikladacham_4(message, grypa, user_id):
    subject = message.text
    subject = subject.lower()
    db_filename = grypa + '.db'
    if os.path.exists(db_filename):

        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f'SELECT "Предмети" FROM Предмети')
        res = cursor.fetchall()
        res = [item for tpl in res for item in tpl]
        if subject == '🔙назад':
            menu_vikladacham(message)
        elif subject in res:
            cursor.execute(f"SELECT Форма_підсумкового_контролю FROM Предмети WHERE Предмети = ?",(subject,))
            key = cursor.fetchall()
            key = [item for tpl in key for item in tpl]
            key = ''.join([str(i) for i in key])
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton('Модуль 1')
            item2 = types.KeyboardButton('Модуль 2')
            item3 = types.KeyboardButton('ІНДЗ')
            item5 = types.KeyboardButton(f'Виставити одразу {key}')
            item4 = types.KeyboardButton('Індивідуальні години')
            close_sub = types.KeyboardButton('Закрити предмет')
            back = types.KeyboardButton('🔙Назад')
            markup.add(back)
            markup.add(item1,item2)
            markup.add(item5)
            markup.add(item3)
            markup.add(item4)
            markup.add(close_sub)
            bot.send_message(message.chat.id, f"Будь ласка, оберіть розділ у предметі '{subject}', який ви хочете відредагувати, закрити або виставити оцінку.\n\n‼️‼️‼️\nЗвертаємо увагу, що функція закриття предмету призведе до втрати можливості редагувати журнал; буде доступний лише режим перегляду.", reply_markup=markup)
            conn.close()
            bot.register_next_step_handler(message, menu_vikladacham_5, db_filename, subject,key)

        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item2 = types.KeyboardButton('Робота з журналами')
            homework = types.KeyboardButton('Додати домашнє')
            item4 = types.KeyboardButton('Оголошення для групи')
            back = types.KeyboardButton('🔙Назад')
            markup.add(back)
            markup.add(item2)
            markup.add(homework)
            markup.add(item4)
            bot.send_message(message.chat.id,"Такого предмету староста не додала в свій журнал",reply_markup=markup)
            bot.register_next_step_handler(message, menu_vikladacham_2)



    else:  # 2_1234
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item2 = types.KeyboardButton('Робота з журналами')
        homework = types.KeyboardButton('Додати домашнє')
        item4 = types.KeyboardButton('Оголошення для групи')
        back = types.KeyboardButton('🔙Назад')
        markup.add(back)
        markup.add(item2)
        markup.add(homework)
        markup.add(item4)
        bot.send_message(message.chat.id, f"Староста групи {grypa} ще не створила журнал зверніть до неї з цим проханням", reply_markup=markup)
        bot.register_next_step_handler(message,menu_vikladacham_2)
def menu_vikladacham_5(message, db_filename,subject,key):
    text = message.text
    if text == 'Модуль 1' or text == 'Модуль 2':
        if text == 'Модуль 1':
            table = '1'
        elif text == 'Модуль 2':
            table = '2'
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        back = types.KeyboardButton('🔙Назад')
        item1 = types.KeyboardButton('Додати оцінку')
        item7 = types.KeyboardButton('Перегляд оцінок')
        item2 = types.KeyboardButton('Ред. назву тему')  #
        item3 = types.KeyboardButton('Додати тему')  #
        item5 = types.KeyboardButton('Закрити модуль') #
        markup.add(back)
        markup.add(item1,item7)
        markup.add(item2, item3)
        markup.add(item5)
        bot.send_message(message.chat.id, f"Оберіть, що саме ви хочете редагувати в журналі предмету {subject}, у розіділі {text}", reply_markup=markup)
        bot.register_next_step_handler(message, menu_vikladacham_add_grate_modul, db_filename, subject, table,key)

    elif text == '🔙Назад':
        menu_vikladacham_3_1(message)


    elif text == 'ІНДЗ':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()

        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()


        students_list = "\n".join([student[0] + ' - 0' for student in students])
        gryp = db_filename.split('.')

        bot.send_message(message.chat.id,f"Ось знизу я надіслав вам список групи {gryp[0]} Ви можете виставити оцінки через - навпроти кожного студента якщо біля якогось студента немає оцінки або він не отримав балів можете сміло ставити 0 або нічого і стерти його із списку вкажіть оцінки за зразком", reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.send_message(message.chat.id,
                         "ПІБ Одногрупника - оцінка\nПІБ Одногрупника - оцінка\nПІБ Одногрупника  - оцінка")
        bot.send_message(message.chat.id, f"<code>{students_list}</code>", parse_mode=ParseMode.HTML)
        bot.register_next_step_handler(message,indz_2, db_filename, subject)


    elif text == f'Виставити одразу {key}':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()

        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()


        students_list = "\n".join([student[0] + ' - 0' for student in students])
        gryp = db_filename.split('.')

        bot.send_message(message.chat.id,f'Ось знизу я надіслав вам список групи {gryp[0]} Ви можете виставити оцінки з {key} через " - " навпроти кожного студента якщо біля якогось студента немає оцінки або він не отримав балів можете сміло ставити 0 або нічого і стерти його із списку вкажіть оцінки за зразком',reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.send_message(message.chat.id,"ПІБ Одногрупника - оцінка\nПІБ Одногрупника - оцінка\nПІБ Одногрупника  - оцінка")
        bot.send_message(message.chat.id,f"Список групи {gryp[0]}:\n <code>{students_list}</code>", parse_mode=ParseMode.HTML)
        bot.register_next_step_handler(message,exam_assessment, db_filename, subject, key)


    elif text == 'Індивідуальні години':
        bot.send_message(message.chat.id, f"Функція в розробці")
        bot.register_next_step_handler(message, menu_vikladacham_5, db_filename, subject, key)
    elif text == 'Закрити предмет':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Закритий_предмет FROM Предмети WHERE Предмети = '{subject}'")
        close = cursor.fetchone()
        result_variable = close[0]

        if result_variable == 'Закритий предмет':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton('Так')
            item2 = types.KeyboardButton('Ні')
            markup.add(item1, item2)
            bot.send_message(message.chat.id,"Предмет вже закритий можливо ви б хотіли його відкрити?)))",reply_markup=markup)
            bot.register_next_step_handler(message, open_subject, db_filename, subject)
        else:
            close_subject(message, db_filename, subject)

    else:
        bot.send_message(message.chat.id, f"Такого варіанту немає")
        bot.register_next_step_handler(message, menu_vikladacham_5, db_filename, subject, key)
def open_subject(message,db_filename,subject):
    text = message.text
    if text == 'Так':
        subject = subject.replace("_", " ")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f'UPDATE Предмети SET Закритий_модуль_1 = NULL WHERE Предмети = ?',(subject,))
        cursor.execute(f'UPDATE Предмети SET Закритий_модуль_2 = NULL WHERE Предмети = ?', (subject,))
        cursor.execute(f'UPDATE Предмети SET Закритий_предмет = NULL WHERE Предмети = ?', (subject,))
        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, f"Предмет {subject} відкритий ")
        menu_vikladacham(message)
    elif text == 'Ні':
        menu_vikladacham(message)
    else:
        bot.send_message(message.chat.id, "Такого варіанту відповіді немає")
        bot.register_next_step_handler(message, open_subject, db_filename, subject)
def close_subject(message,db_filename,subject):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton('Так🥶')
    item2 = types.KeyboardButton('Ні🥵')
    markup.add(item1, item2)
    bot.send_message(message.chat.id,
                     f"Ви обрали функцію закриття предмету {subject}. Чи ви впевнені, що хочете позбавити старост статусу редагування цього предмету? Будь ласка, оберіть один із варіантів відповіді нижче.",
                     reply_markup=markup)
    bot.register_next_step_handler(message, close_subject_2, db_filename, subject)
def close_subject_2(message,db_filename,subject):
    text = message.text
    if text == 'Так🥶':
        subject = subject.replace("_", " ")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        a = 'Закритий предмет'
        a1 = 'Закритий модуль 1'
        a2 = 'Закритий модуль 2'
        cursor.execute(f'UPDATE Предмети SET Закритий_модуль_1 = ? WHERE Предмети = ?',(a1,subject))
        cursor.execute(f'UPDATE Предмети SET Закритий_модуль_2 = ? WHERE Предмети = ?', (a2, subject))
        cursor.execute(f'UPDATE Предмети SET Закритий_предмет = ? WHERE Предмети = ?', (a, subject))
        conn.commit()
        conn.close()
        bot.send_message(message.chat.id,"Предмет закритий😒")
        time.sleep(2)
        bot.send_message(message.chat.id, "Надіюсь студенти не будуть на вас дутися😋")
        time.sleep(1)

        menu_vikladacham(message)
    elif text == 'Ні🥵':
        menu_vikladacham(message)
    else:
        bot.send_message(message.chat.id, "Такого варіанту відповіді немає")
        bot.register_next_step_handler(message, close_subject_2, db_filename, subject)
def exam_assessment(message, db_filename, subject, key):
    text = message.text
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    rows = text.split("\n")
    subject = subject.replace(" ","_")
    split = []
    splitnot = []
    for row in rows:
        student_data = row.split(" - ")
        if len(student_data) == 2:
            split.append(row)

        else:
            splitnot.append(row)

    if len(splitnot) == 0:
        for row in split:
            student_data = row.split(" - ")
            name, grade = student_data
            cursor.execute(f"UPDATE {subject}_3 SET [{key}] = ? WHERE Студенти = ?",(grade, name))
            cursor.execute(f"UPDATE {subject}_Студенти SET [{key}] = ? WHERE Студенти = ?",(grade, name))

        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, "Дані успішно додано до бази даних.")
        table_name = f'{subject}_1'
        Jurnal_1_5.jurnal1_5_dodavanna_ocinok(message, db_filename, table_name, subject)


    elif len(splitnot) > 0:
        list = tuple(splitnot)
        message_text = "\n\n".join(list)
        bot.send_message(message.chat.id,
                         f"{message_text} Ось ці рядки я не зміг розпізнати, будь ласка надішліть ще раз і правильно за таким зразком\n\nПІБ - Оцінка\nПІБ - Оцінка")
        bot.register_next_step_handler(message,exam_assessment, db_filename, subject, key)
def menu_vikladacham_add_grate_modul(message, db_filename,subject, table,key):
    text = message.text
    if text.startswith('/'):
        bot.send_message(message.chat.id, "Ви ввели команду, будь ласка оберіть розділ, яким хочете працювати")
        bot.register_next_step_handler(message, menu_vikladacham_add_grate_modul, db_filename, subject, table)
    elif text ==  '🔙Назад':
        message_handler_start(message)

    elif text == "Додати оцінку":
        subject = subject.replace(" ","_")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        # Формуємо назву таблиці у форматі {subject}_{table}
        table_name = f'{subject}_{table}'
        # Отримуємо назви стовпців таблиці
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if column[1] != 'Студенти' and column[1] != 'Н']
        # Створюємо клавіатуру з кнопками зі списку column_names
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        for column_name in column_names:
            markup.add(column_name)
        # Надсилаємо повідомлення з клавіатурою
        bot.send_message(message.chat.id, "Оберіть тему в яку хочете внести оцінки:", reply_markup=markup)
        bot.register_next_step_handler(message, menu_vikladacham_add_grate_1_modul, db_filename, subject, table,column_names)

    elif text == 'Перегляд оцінок':

        subject = subject.replace(" ", "_")
        table_name = f'{subject}_{table}'
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('1')
        item2 = types.KeyboardButton('2')
        markup.add(item1)
        markup.add(item2)
        bot.send_message(message.chat.id,"Оберіть функцію:\n1 - переглянути оцінки окремого студента\n2 - перегляд оцінок всієї групи із окремої теми", reply_markup=markup)
        bot.register_next_step_handler(message,menu_vikladacham_look_grate_1,db_filename, table_name, subject)

    elif text == 'Ред. назву тему':
        subject = subject.replace(" ","_")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        # Формуємо назву таблиці у форматі {subject}_{table}
        table_name = f'{subject}_{table}'

        # Отримуємо назви стовпців таблиці
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if
                        column[1] != 'Студенти' and column[1] != 'модуль_1' and column[1] != 'модуль_2' and column[
                            1] != 'Н']

        # Створюємо клавіатуру з кнопками зі списку column_names
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        for column_name in column_names:
            markup.add(column_name)


        # Надсилаємо повідомлення з клавіатурою
        user_grypa = db_filename.split(".")
        user_grypa = user_grypa[0]
        bot.send_message(message.chat.id, "Оберіть тему, яку ви хочете відредагувати:", reply_markup=markup)
        bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_1, db_filename, user_grypa, subject, table, column_names)

        conn.close()


    elif text == 'Додати тему':
        subject = subject.replace(" ", "_")
        bot.send_message(message.chat.id, "Надішліть назву нової теми для предмету",
                         reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy, db_filename, subject, table)



    elif text == 'Закрити модуль':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Закритий_модуль_{table} FROM Предмети WHERE Предмети = '{subject}'")
        close = cursor.fetchone()
        result_variable = close[0]

        if result_variable == f'Закритий модуль {table}':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton('Так')
            item2 = types.KeyboardButton('Ні')
            markup.add(item1, item2)
            bot.send_message(message.chat.id,f"Модуль {table} вже закритий можливо ви б хотіли його відкрити?)))",reply_markup=markup)
            bot.register_next_step_handler(message, open_module, db_filename, subject, table)
        else:
            close_module_1(message, db_filename, subject,table)
    else:
        bot.send_message(message.chat.id, "Такого варіанту немає")
        bot.register_next_step_handler(message, menu_vikladacham_add_grate_modul, db_filename, subject, table, key)



def open_module(message, db_filename, subject,table):
    text = message.text
    if text == 'Так':
        subject = subject.replace("_", " ")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()


        cursor.execute(f'UPDATE Предмети SET Закритий_модуль_{table} = NULL WHERE Предмети = ?', (subject,))

        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, f"Модуль {table} відкритий для старост")
        menu_vikladacham(message)
    elif text == 'Ні':
        menu_vikladacham(message)
    else:
        bot.send_message(message.chat.id, "Такого варіанту відповіді немає")
        bot.register_next_step_handler(message, open_subject, db_filename, subject)
def close_module_1(message, db_filename, subject,table):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton('Так🥶')
    item2 = types.KeyboardButton('Ні🥵')
    markup.add(item1, item2)
    bot.send_message(message.chat.id,
                     f"Ви обрали функцію закриття Модуля {table} для предмету {subject}. Чи ви впевнені, що хочете позбавити старост статусу редагування цього модуля? Будь ласка, оберіть один із варіантів відповіді нижче.",
                     reply_markup=markup)
    bot.register_next_step_handler(message, close_module, db_filename, subject, table)
def close_module(message,db_filename, subject, table):
    text = message.text
    if text == 'Так🥶':
        subject = subject.replace("_", " ")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        a = f'Закритий модуль {table}'
        cursor.execute(f'UPDATE Предмети SET Закритий_модуль_{table} = ? WHERE Предмети = ?', (a, subject))
        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, f"ви відкрили Модуль {table} із предмету {subject}")
        menu_vikladacham(message)
    elif text == 'Ні🥵':
        menu_vikladacham(message)
    else:
        bot.send_message(message.chat.id, "Такого варіанту відповіді немає")
        bot.register_next_step_handler(message,close_module, db_filename, subject, table)

def menu_vikladacham_look_grate_1(message,db_filename, table_name, subject):

    text = message.text
    if text == "2":
        menu_vikladacham_look_grate_2_1(message, db_filename, table_name, subject)
    elif text == '1':
        bot.send_message(message.chat.id, "Функція скоро буде")
def menu_vikladacham_look_grate_2_1(message, db_filename, table_name, subject):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = cursor.fetchall()
    column_names = [column[1] for column in columns if column[1] != 'Студенти' and column[1] != 'Н']

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('🔙Назад'))
    for column_name in column_names:
        markup.add(column_name)

    bot.send_message(message.chat.id,"Оберіть тему, з якої ви хочете переглянути оцінки своїх студентів або зразу з цілого модуля:",reply_markup=markup)
    bot.register_next_step_handler(message, menu_vikladacham_look_grate_2_2, db_filename, table_name,column_names, subject)
def menu_vikladacham_look_grate_2_2(message, db_filename, table_name,column_names, subject):
    text = message.text
    if text == '🔙Назад':
        menu_vikladacham(message)
    elif text not in column_names:
        bot.send_message(message.chat.id, "Ви вибрали не вірну тему оберіть ще раз")
        bot.register_next_step_handler(message, menu_vikladacham_look_grate_2_2, db_filename,table_name,column_names,subject)
    else:

        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()
        # Формуємо повідомлення зі списком студентів
        students_list = "\n".join([student[0] + ' - ' for student in students])

        cursor.execute(f"SELECT [{text}] FROM {table_name} ")
        ocinky = cursor.fetchall()
        ocinky = ocinky[2:]
        ocinky = [item for tpl in ocinky for item in tpl]
        ocinky2 = []
        for i in ocinky:
            if i == None:
                ocinky2.append("0")
            else:
                ocinky2.append(i)

        results = []
        students_list1 = students_list.split("\n")
        for k, i in enumerate(students_list1):
            results.append(i + ocinky2[k])

        gem = "\n".join([row for row in results])
        subject = subject.replace("_", " ")
        bot.send_message(message.chat.id, f"Ось оцінки із теми {text} із предмету {subject}:\n{gem}")
        bot.register_next_step_handler(message, menu_vikladacham_look_grate_2_2, db_filename, table_name, column_names,subject)
def menu_vikladacham_add_grate_1_modul(message, db_filename,subject, table,column_names):
    tema = message.text
    if tema == '🔙Назад':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Форма_підсумкового_контролю FROM Предмети WHERE Предмети = ?", (subject,))
        key = cursor.fetchall()
        key = [item for tpl in key for item in tpl]
        key = ''.join([str(i) for i in key])
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Модуль 1')
        item2 = types.KeyboardButton('Модуль 2')
        item3 = types.KeyboardButton('ІНДЗ')
        item5 = types.KeyboardButton(f'Виставити одразу {key}')
        item4 = types.KeyboardButton('Індивідуальні години')
        close_sub = types.KeyboardButton('Закрити предмет')
        back = types.KeyboardButton('🔙Назад')
        markup.add(back)
        markup.add(item1, item2)
        markup.add(item5)
        markup.add(item3)
        markup.add(item4)
        markup.add(close_sub)
        bot.send_message(message.chat.id,
                         f"Оберіть розділ в предметі {subject}, який хочете редагувати або додати оцінку або закрити предмет(Функція закриття предмету забирає змогу в старост редагувати журнал їм буде дозволений тільки перегляд)",
                         reply_markup=markup)
        bot.register_next_step_handler(message, menu_vikladacham_5, db_filename, subject, key)

    elif tema in column_names:
        bot.send_message(message.chat.id,
                         "Будь ласка надішліть мені список студентів групи їхнє повне ім\'я.\nЗа таким зразком\nТакож я надішлю вам групи, для зручнішого виставлення оцінок")
        bot.send_message(message.chat.id,
                         "ПІБ(Одногрупника) - оцінка\nПІБ(Одногрупника) - оцінка\nПІБ(Одногрупника) - оцінка")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()

        # Витягуємо всі значення зі стовпця "Студенти" таблиці "STUDENTY"
        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()

        # Формуємо повідомлення зі списком студентів
        students_list = "\n".join([student[0] + ' - 0' for student in students])
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        # Надсилаємо повідомлення зі списком студентів у бота
        bot.send_message(message.chat.id, f"<code>{students_list}</code>", parse_mode=ParseMode.HTML, reply_markup=markup)
        conn.close()
        bot.register_next_step_handler(message, menu_vickladacham_add_grate_2, db_filename, subject, table, tema)
    else:
        bot.send_message(message.chat.id,"Немає такого варіанту відповіді")
        bot.register_next_step_handler(message, menu_vikladacham_add_grate_1_modul, db_filename, subject, table,column_names)
def menu_vickladacham_add_grate_2(message, db_filename, subject, table, tema):
    text = message.text
    if text == '🔙Назад':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Форма_підсумкового_контролю FROM Предмети WHERE Предмети = ?", (subject,))
        key = cursor.fetchall()
        key = [item for tpl in key for item in tpl]
        key = ''.join([str(i) for i in key])
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Модуль 1')
        item2 = types.KeyboardButton('Модуль 2')
        item3 = types.KeyboardButton('ІНДЗ')
        item5 = types.KeyboardButton(f'Виставити одразу {key}')
        item4 = types.KeyboardButton('Індивідуальні години')
        close_sub = types.KeyboardButton('Закрити предмет')
        back = types.KeyboardButton('🔙Назад')
        markup.add(back)
        markup.add(item1, item2)
        markup.add(item5)
        markup.add(item3)
        markup.add(item4)
        markup.add(close_sub)
        bot.send_message(message.chat.id,
                         f"Оберіть розділ в предметі {subject}, який хочете редагувати або додати оцінку або закрити предмет(Функція закриття предмету забирає змогу в старост редагувати журнал їм буде дозволений тільки перегляд)",
                         reply_markup=markup)
        bot.register_next_step_handler(message, menu_vikladacham_5, db_filename, subject, key)
    else:
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        rows = text.split("\n")
        table_name = f'{subject}_{table}'
        split = []
        splitnot = []
        for row in rows:
            student_data = row.split(" - ")
            if len(student_data) == 2:
                split.append(row)

            else:
                splitnot.append(row)

        if len(splitnot) == 0:
            for row in split:
                student_data = row.split(" - ")
                name, grade = student_data
                cursor.execute(
                    f"UPDATE {table_name} SET [{tema}] = ? WHERE Студенти = ?",
                    (grade, name))

            conn.commit()
            conn.close()
            bot.send_message(message.chat.id, "Дані успішно додано до бази даних.")
            Jurnal_1_5.jurnal1_5_dodavanna_ocinok(message, db_filename, table_name, subject)

        elif len(splitnot) > 0:
            list = tuple(splitnot)
            message_text = "\n\n".join(list)
            bot.send_message(message.chat.id,f"{message_text} Ось ці рядки я не зміг розпізнати, будь ласка надішліть ще раз і правильно за таким зразком\n\nПІБ - Оцінка\nПІБ - Оцінка", reply_markup=telebot.types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message, menu_vickladacham_add_grate_2, db_filename, subject, table, tema)





@bot.message_handler(commands=['menu'])
def message_handler_start(message):
    user_id = message.from_user.id
    user_id = str(user_id)
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
    user_rol = cursor.fetchone()

    cursor.execute(f"SELECT ID FROM Диспетчери ")
    user_des = cursor.fetchall()
    user_des = [item for tpl in user_des for item in tpl]

    if user_rol:
        user_rol = user_rol[0]
        if user_rol == 'староста':
            if user_rol == 'староста' and user_id in user_des:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton('📜Профіль')
                item2 = types.KeyboardButton('✍️Розклад пар')
                item3 = types.KeyboardButton('⚠️Контакти викладачів')
                jurnal = types.KeyboardButton('📔Журнал')
                item4 = types.KeyboardButton('🤓Старостам')
                item5 = types.KeyboardButton('📝Домашка')
                item5_6 = types.KeyboardButton('🍺Підтримка проекту🌯')
                itemdes = types.KeyboardButton('Диспетчерам')
                markup.add(item1, item2, item3,jurnal, item4, item5)

                markup.add(item5_6)
                markup.add(itemdes)
                bot.send_message(message.chat.id, "👇".format(message.from_user), reply_markup=markup)
            else:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton('📜Профіль')
                item2 = types.KeyboardButton('✍️Розклад пар')
                item3 = types.KeyboardButton('⚠️Контакти викладачів')
                jurnal = types.KeyboardButton('📔Журнал')
                item4 = types.KeyboardButton('🤓Старостам')
                item5 = types.KeyboardButton('📝Домашка')
                item5_6 = types.KeyboardButton('🍺Підтримка проекту🌯')
                markup.add(item1, item2, item3, jurnal, item4, item5)
                markup.add(item5_6)
                bot.send_message(message.chat.id, "👇".format(message.from_user), reply_markup=markup)

        elif user_rol == 'викладач':
            if user_rol == 'викладач' and user_id in user_des:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton('📜Профіль')
                item2 = types.KeyboardButton('✍️Розклад пар')
                item3 = types.KeyboardButton('⚠️Контакти викладачів')
                item4 = types.KeyboardButton('Викладачам')
                item5_6 = types.KeyboardButton('🍺Підтримка проекту🌯')
                itemdes = types.KeyboardButton('Диспетчерам')
                markup.add(item1, item2, item3, item4)
                markup.add(item5_6)
                markup.add(itemdes)
                bot.send_message(message.chat.id, "👇".format(message.from_user), reply_markup=markup)
            else:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton('📜Профіль')
                item2 = types.KeyboardButton('✍️Розклад пар')
                item3 = types.KeyboardButton('⚠️Контакти викладачів')
                item4 = types.KeyboardButton('Викладачам')
                item5_6 = types.KeyboardButton('🍺Підтримка проекту🌯')
                markup.add(item1, item2, item3, item4)
                markup.add(item5_6)
                bot.send_message(message.chat.id, "👇".format(message.from_user), reply_markup=markup)
        elif user_rol == 'студент':
            if  user_rol == 'студент' and user_id in user_des:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton('📜Профіль')
                item2 = types.KeyboardButton('✍️Розклад пар')
                item3 = types.KeyboardButton('⚠️Контакти викладачів')
                jurnal = types.KeyboardButton('📔Журнал')
                item5 = types.KeyboardButton('📝Домашка')
                item5_6 = types.KeyboardButton('🍺Підтримка проекту🌯')
                itemdes = types.KeyboardButton('Диспетчерам')
                markup.add(item1, item2, item3,jurnal, item5)
                markup.add(item5_6)
                markup.add(itemdes)
                bot.send_message(message.chat.id, "👇".format(message.from_user), reply_markup=markup)
            else:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item1 = types.KeyboardButton('📜Профіль')
                item2 = types.KeyboardButton('✍️Розклад пар')
                item3 = types.KeyboardButton('⚠️Контакти викладачів')
                jurnal = types.KeyboardButton('📔Журнал')
                item5 = types.KeyboardButton('📝Домашка')
                item5_6 = types.KeyboardButton('🍺Підтримка проекту🌯')
                markup.add(item1, item2, item3, jurnal, item5)
                markup.add(item5_6)
                bot.send_message(message.chat.id, "👇".format(message.from_user), reply_markup=markup)

def support_project(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_menu = types.KeyboardButton('Інформація про розробників')
    info = types.KeyboardButton('Інформація про підтримку')
    info2 = types.KeyboardButton('Донатик')
    info3 = types.KeyboardButton('Додати відгук')
    back = types.KeyboardButton('🔙Назад')
    markup.add(back)
    markup.add(item_menu)
    markup.add(info)
    markup.add(info2)
    markup.add(info3)
    bot.send_message(message.chat.id, "Будемо раді і ситі, якщо ви підтримаєте проект теплими словами та\або матеріально)🥹", reply_markup=markup)
    bot.register_next_step_handler(message,support_project2)
def support_project2(message):
    text = message.text
    if text == 'Інформація про підтримку':
        bot.send_message(message.chat.id,"Привіт, ми збираємо кошти на те щоб проводити розробку та вдосконалення нашого загального бота під пиво та\або чай з печивом, памʼятайте донат не є обов'язковим,але нам буде дуже приємно якщо ви підтримаєте нас копійкою, але першочергово кидайте донат на ЗСУ, дякуємо всім, ну і звісно нагадуємо що маленьких донатів не буває)")
        bot.register_next_step_handler(message, support_project2)
    elif message.text == '🔙Назад':
        message_handler_start(message)
    elif text == 'Донатик':
        bot.send_message(message.chat.id,"На пиво та шавуху)\n\nПосилання на банку\nhttps://send.monobank.ua/jar/7SDhjdhEQd\n💳Номер картки банки\n 5375 4112 0849 8779")
        bot.register_next_step_handler(message, support_project2)
    elif text == 'Додати відгук':
        bot.send_message(message.chat.id, 'Це наша невеличка книга скарг та пропозицій, з великим заохоченням прочитаємо усі ваші відгуки, та дякуємо за підтримку нашого проекту словом)\n P.s. дякуй в кешеню не покладеш', reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, response)
    elif text == 'Інформація про розробників':
        bot.send_message(message.chat.id, '@yura_krykh\n@taraszubik9\n@mamyn_synok\n\nІ також наші перші тестери які помогли не тільки цим: \n@s_melnyk18\n@Da_rrr_lin_g')
        bot.register_next_step_handler(message, support_project2)
    else:
        bot.send_message(message.chat.id, 'Немає такого варіанту')
        bot.register_next_step_handler(message, support_project2)
def response(message):
    text = message.text
    id = message.chat.id
    conn = sqlite3.connect(f"users.db")
    cursor = conn.cursor()
    cursor.execute('INSERT INTO "Відгуки" (id, text) VALUES (?, ?)', (id, text))
    conn.commit()
    conn.close()
    bot.send_message(message.chat.id, 'Дякую за відгук)\nP.s. В цьому випадку в кишеню нічого класти не треба😉')
    message_handler_start(message)



@bot.message_handler(content_types=['text'])
def bot_message(message):
    if message.chat.type == 'private':
        if message.text == '🍺Підтримка проекту🌯':
            support_project(message)

        elif message.text == 'Диспетчерам':
            user_id = message.from_user.id
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            cursor.execute(f"SELECT ID FROM Диспетчери ")
            user_des = cursor.fetchall()
            user_des = [item for tpl in user_des for item in tpl]

            if str(user_id) in user_des:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                markup.add(types.KeyboardButton('🔙Назад'))
                markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
                markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
                markup.add(types.KeyboardButton("Додати диспетчера"), types.KeyboardButton("Додати навчальні плани"))
                markup.add(types.KeyboardButton("Кількість користувачів"), types.KeyboardButton("Редагувати студента"))

                bot.send_message(message.chat.id, "Меню диспетчера", reply_markup=markup)
                bot.register_next_step_handler(message, Menu_dess.menu_desp)

        elif message.text == 'Додати домашку':
            bot.send_message(message.chat.id, "Функція в розробці")
            bot.register_next_step_handler(message,bot_message)

        elif message.text == '🤓Старостам':
            user_id = message.from_user.id
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            cursor.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
            user_rol = cursor.fetchone()
            if user_rol:
                user_rol = user_rol[0]
                if user_rol == 'староста':
                    menu_starostam(message)
                else:
                    bot.send_message(message.chat.id, "Ви не є старостою, ви не можете користуватися цим меню)")
                    message_handler_start(message)
            else:
                bot.send_message(message.chat.id, "Користувача не знайдено в базі даних")
                message_handler_start(message)

        elif message.text == 'Викладачам':
            chet_teacer(message)

        elif message.text == 'Редагувати розклад':
            user_id = message.from_user.id
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            cursor.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
            user_rol = cursor.fetchone()
            if user_rol:
                user_rol = user_rol[0]
                if user_rol == 'староста':
                    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    item1 = types.KeyboardButton('Понеділок')
                    item2 = types.KeyboardButton('Вівторок')
                    item3 =types.KeyboardButton('Середа')
                    item4 = types.KeyboardButton('Четвер')
                    item5 = types.KeyboardButton('П\'ятниця')
                    back =types.KeyboardButton('🔙Назад')
                    keyboard.add(back)
                    keyboard.add(item1, item2, item3, item4, item5)

                    bot.send_message(message.chat.id,"Оберіть день, у який ви внесете корекцію", reply_markup=keyboard)
                    bot.register_next_step_handler(message, redaguvanna,user_id)
                else:
                    bot.send_message(message.chat.id, "Ви не є старостою, ви не можете користуватися цим меню)")
                    message_handler_start(message)

        elif message.text == '📝Домашка':
            bot.send_message(message.chat.id, "Функція в розробці")
            bot.register_next_step_handler(message, bot_message)

        elif message.text == '✍️Розклад пар':
            user_id = message.chat.id
            connect = sqlite3.connect('users.db')
            cursor = connect.cursor()
            cursor.execute("SELECT grypa FROM login_id WHERE id = ?", (user_id,))
            rows = cursor.fetchone()
            user_grypa = rows[0]
            cursor.execute("SELECT Групи FROM Групи")
            rows = cursor.fetchall()
            gryps = [row[0] for row in rows]
            if user_grypa.replace("_","-") in gryps:
                Rozklad.rozklad_par_0_1(message, user_id, user_grypa)
            else:
                Rozklad.rozklad_par_0(message, user_id)

        elif message.text == "🛠Редагування профілю":
            user_id = message.chat.id
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            bot.send_message(message.chat.id, "Надішліть, що саме хочете відредагувати або просто надішліть одразу відредагований профіл")
            bot.register_next_step_handler(message,edit_profile, user_id)

        elif message.text == '📜Профіль':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item = types.KeyboardButton('🛠Редагування профілю')
            back = types.KeyboardButton('🔙Назад')
            markup.add(back)
            markup.add(item)

            user_id = message.chat.id
            user_data = get_user_data(user_id)
            if user_data is not None:
                grypa, email, first_last = user_data
                profile_info = f"📜Профіль\n📚Група: {grypa}\n✉️Email: {email}\n👨‍🎓ПІБ: {first_last}"
                bot.send_message(message.chat.id, profile_info, reply_markup=markup)
            else:
                bot.send_message(message.chat.id, "Вас не знайдено(\nЗвернись до /support!")

        elif message.text == '⚠️Контакти викладачів':
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()

            cursor.execute("SELECT Викладач FROM Teachers")
            teachers = cursor.fetchall()


            keyboard = []
            for i in range(0, len(teachers), 3):
                row = []
                for j in range(i, min(i + 3, len(teachers))):
                    teacher_name = teachers[j][0]
                    row.append(InlineKeyboardButton(teacher_name, callback_data=teacher_name))
                keyboard.append(row)


            reply_markup = InlineKeyboardMarkup(keyboard)
            reply_markup = json.dumps(reply_markup.to_dict())

            conn.close()
            bot.send_message(message.chat.id, "Оберіть викладача за прізвищем якого шукаєте:",reply_markup=reply_markup)

        elif message.text == "Робота з журналом":
            user_id = message.from_user.id
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            cursor.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
            user_rol = cursor.fetchone()
            if user_rol:
                user_rol = user_rol[0]
                if user_rol == 'староста':
                    jurnal1(message)
                else:
                    bot.send_message(message.chat.id, "Ви не є старостою, ви не можете користуватися цим меню)")
                    message_handler_start(message)
            else:
                bot.send_message(message.chat.id, "Користувача не знайдено в базі даних")
                message_handler_start(message)

        elif message.text == '🔙Назад':
            message_handler_start(message)
        elif message.text == '📔Журнал':
            user_id = message.chat.id
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            cursor.execute(f"SELECT grypa, first_last FROM login_id WHERE id = {user_id}")
            row = cursor.fetchone()
            if row is not None:
                user_grypa = row[0]
                first_last = row[1]
            else:
                # Обробка випадку, коли користувач не знайдений
                user_grypa = None
                first_last = None
            user_grypa = user_grypa.upper().replace('_', '-')

            db_filename = user_grypa + '.db'
            if os.path.exists(db_filename):  # 1_1234
                bot.send_message(message.chat.id, f"Я знайшов твій журнал 🙂")
                pereglad_ocinok(message, db_filename,first_last )


            else:  # 2_1234
                bot.send_message(message.chat.id,f"Схоже ваша староста ще не створила журнал для вашої групи {user_grypa}🫠\nВи можете її вічливо попросити це зробити")

        elif message.text == 'Оголошення для групи':
            ogoloshennya_grypa(message)

        else:
            text = message.text
            bot_message1(message,text)


def edit_profile(message, user_id):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f"SELECT ID FROM Диспетчери")
    results = cursor.fetchall()
    results = [item for tpl in results for item in tpl]
    for i in results:
        bot.send_message(chat_id=i,text= f"Користувач під айді {user_id} хоче відредагувати профіль, ось так...")
        bot.forward_message(chat_id=i, from_chat_id=message.chat.id, message_id=message.message_id)

    bot.send_message(user_id, "Повідомлення надіслане адміністрації очікуйте схвалення на редагування профілю")
    message_handler_start(message)







class Menu_dess:
    def __init__(self,bot):
        self.bot = bot
    @staticmethod
    def menu_desp(message):
        text = message.text
        if text == '🔙Назад':
            message_handler_start(message)
        elif text == 'Редагувати розклад':
            connect = sqlite3.connect('users.db')
            cursor = connect.cursor()
            cursor.execute("SELECT Групи FROM Групи")
            rows = cursor.fetchall()
            gryps = [row[0] for row in rows]
            columns = 3
            gryps_per_column = (len(gryps) + columns - 1) // columns
            gryps_divided = [gryps[i:i + gryps_per_column] for i in range(0, len(gryps), gryps_per_column)]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            # Додавання груп до розмітки по 3 в кожному рядку
            for gryp_column in gryps_divided:
                markup.add(*gryp_column)

            connect.close()
            bot.send_message(message.chat.id, 'Оберіть групу в якій хочете редагувати розклад:', reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.red_rozklad, gryps)

        elif text == 'Перегляд оцінок':
            connect = sqlite3.connect('users.db')
            cursor = connect.cursor()
            cursor.execute("SELECT Групи FROM Групи")
            rows = cursor.fetchall()
            gryps = [row[0] for row in rows]
            columns = 3
            gryps_per_column = (len(gryps) + columns - 1) // columns
            gryps_divided = [gryps[i:i + gryps_per_column] for i in range(0, len(gryps), gryps_per_column)]

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            # Додавання груп до розмітки по 3 в кожному рядку
            for gryp_column in gryps_divided:
                markup.add(*gryp_column)

            connect.close()
            bot.send_message(message.chat.id, 'Оберіть групу в якій хочете переглянути оцінки:', reply_markup=markup)
            bot.register_next_step_handler(message,Menu_dess.look_grate_stud)

        elif text == 'Обновлення пошт':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('Схвалити'))
            markup.add(types.KeyboardButton('Відмінити'))
            bot.send_message(message.chat.id, "Ви точно хочете додати нові фізматівські пошти у базу даних?", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.add_new_email)

        elif text == 'Редагування списку груп':
            bot.send_message(message.chat.id, "Степан ну почекай я не всьо встиг")
            bot.register_next_step_handler(message, Menu_dess.menu_desp)

        elif text == "Додати диспетчера":
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            bot.send_message(message.chat.id, "Надішліть мені айді користувача якого хочете зробити новим диспетчером, не забудьте йому розповісти про його обов\'язки", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.add_desp)


        elif text == "Додати навчальні плани":
            bot.send_message(message.chat.id, "Стьопа скоро буде")
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        elif text == "Кількість користувачів":
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            cursor.execute(f'SELECT id FROM login_id')
            users = cursor.fetchall()
            users = len(users)
            bot.send_message(message.chat.id, f"Кількість користувачів {users}")
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        elif text == "Редагувати студента":
            bot.send_message(message.chat.id, "Стьопа скоро буде")
            bot.register_next_step_handler(message, Menu_dess.menu_desp)




    @staticmethod
    def add_desp(message):
        text = message.text
        if text == '🔙Назад':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
            markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
            markup.add(types.KeyboardButton("Додати диспетчера"), types.KeyboardButton("Додати навчальні плани"))
            markup.add(types.KeyboardButton("Кількість користувачів"), types.KeyboardButton("Редагувати студента"))
            bot.send_message(message.chat.id, "Меню диспетчера", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        else:
            if text.isdigit():
                conn = sqlite3.connect('users.db')
                cursor = conn.cursor()
                cursor.execute("SELECT ID FROM Диспетчери")
                des = cursor.fetchall()
                des = [row[0] for row in des]
                if text in des:
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                    markup.add(types.KeyboardButton('🔙Назад'))
                    markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
                    markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
                    markup.add(types.KeyboardButton("Додати диспетчера"),types.KeyboardButton("Додати навчальні плани"))
                    markup.add(types.KeyboardButton("Кількість користувачів"),types.KeyboardButton("Редагувати студента"))
                    bot.send_message(message.chat.id, "Такий диспетчер вже існує", reply_markup=markup)
                    bot.register_next_step_handler(message, Menu_dess.menu_desp)
                else:
                    cursor.execute("SELECT id FROM login_id")
                    all_user = cursor.fetchall()
                    all_user = [row[0] for row in all_user]
                    if text not in all_user:
                        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                        markup.add(types.KeyboardButton('🔙Назад'))
                        markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
                        markup.add(types.KeyboardButton('Обновлення пошт'),types.KeyboardButton('Редагування списку груп'))
                        markup.add(types.KeyboardButton("Додати диспетчера"),types.KeyboardButton("Додати навчальні плани"))
                        markup.add(types.KeyboardButton("Кількість користувачів"),types.KeyboardButton("Редагувати студента"))
                        bot.send_message(message.chat.id,"Користувач під таким айді не зареєстрований у базі данних, нажаль я його додати не зможу", reply_markup=markup)
                        bot.register_next_step_handler(message, Menu_dess.menu_desp)
                    else:
                        cursor.execute(f"INSERT INTO Диспетчери (ID) VALUES ({text})")
                        conn.commit()
                        conn.close()
                        bot.send_message(message.chat.id, f"Диспетчера {text} успішно додано")
                        message_handler_start(message)
            else:
                bot.send_message(message.chat.id, "Айді диспетчера має складатися тільки із цифр, введіть ше раз айді")
                bot.register_next_step_handler(message, Menu_dess.add_desp)
    @staticmethod
    def look_grate_stud(message):
        text = message.text
        db_filename = text + ".db"
        user_grypa = text
        if os.path.exists(db_filename):
            conn = sqlite3.connect(f'{db_filename}')
            cursor = conn.cursor()
            cursor.execute('SELECT Предмети FROM Предмети')
            result = cursor.fetchall()

            keyboard = []
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for row in result:
                subject = row[0]
                keyboard.append(subject)
                markup.add(types.KeyboardButton(subject))

            conn.close()
            bot.send_message(message.chat.id, f"Оберіть предмет в групі {user_grypa}, з якого хочете переглянути оцінки", reply_markup=markup)
            bot.register_next_step_handler(message, Create.create_table, db_filename, user_grypa)
        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
            markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
            markup.add(types.KeyboardButton("Додати диспетчера"), types.KeyboardButton("Додати навчальні плани"))
            markup.add(types.KeyboardButton("Кількість користувачів"), types.KeyboardButton("Редагувати студента"))
            bot.send_message(message.chat.id,f"Староста групи {user_grypa}, ще не створила журнал, але ви можете попросити її це зробити", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
    @staticmethod
    def add_new_email(message):
        text = message.text
        if text == 'Схвалити':
            bot.send_message(message.chat.id, "Надішліть мені список пошт максимум 25 за 1 раз за зразком нумерація не потрібна\n\n Зразок:\nwwww_ww@fizmat.tnpu.edu.ua - ПІБ нового студента\nwwww_ww@fizmat.tnpu.edu.ua - ПІБ нового студента\nwwww_ww@fizmat.tnpu.edu.ua - ПІБ нового студента\n",reply_markup=telebot.types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message,Menu_dess.add_new_email_2)

        elif text == 'Відмінити':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
            markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
            markup.add(types.KeyboardButton("Додати диспетчера"), types.KeyboardButton("Додати навчальні плани"))
            markup.add(types.KeyboardButton("Кількість користувачів"), types.KeyboardButton("Редагувати студента"))

            bot.send_message(message.chat.id, "Меню диспетчера", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        else:
            bot.send_message(message.chat.id, "Такого варіанту немає")
            bot.register_next_step_handler(message, Menu_dess.add_new_email)
    @staticmethod
    def add_new_email_2(message):
        text = message.text
        split = text.split("\n")
        if len(split)>25:
            bot.send_message(message.chat.id,"Ви надіслали забагато пошт <b>максимальна кількість за 1 раз 25 штук</b>‼️‼️\nНадішліть ще раз",parse_mode=ParseMode.HTML)
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        else:
            email = []
            for i in split:
                k = i.split(" - ")
                email.append(k)

            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            cursor.execute("SELECT Email_Address FROM Email_Base")
            rows = cursor.fetchall()
            emails = [row[0] for row in rows]

            inemailbase = []
            notin = ""
            for i in email:
                poshta, pib = i
                if poshta in emails:
                    inemailbase.append(poshta)

            notin = ",\n".join(inemailbase)
            if len(inemailbase) > 0:
                bot.send_message(message.chat.id,"ВИ надіслали пошти, які вже входять у базу данних\nНадішліть ще раз і постарайцтеся не помилятися",parse_mode=ParseMode.HTML)
                bot.register_next_step_handler(message, Menu_dess.menu_desp)

            else:
                for i in email:
                    poshta, pib = i
                    cursor.execute(f'INSERT INTO "Email_Base" (Email_Address, ПІП) VALUES (?,?)', (poshta, pib,))
                conn.commit()
                conn.close()
                bot.send_message(message.chat.id,"Пошти успішно додані в базу данних✅")
                message_handler_start(message)
    @staticmethod
    def red_rozklad(message, gryps):
        text = message.text
        if text == '🔙Назад':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
            markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
            markup.add(types.KeyboardButton("Додати диспетчера"), types.KeyboardButton("Додати навчальні плани"))
            markup.add(types.KeyboardButton("Кількість користувачів"), types.KeyboardButton("Редагувати студента"))
            bot.send_message(message.chat.id, "Меню диспетчера", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        elif text in gryps:
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton('Понеділок')
            item2 = types.KeyboardButton('Вівторок')
            item3 = types.KeyboardButton('Середа')
            item4 = types.KeyboardButton('Четвер')
            item5 = types.KeyboardButton('П\'ятниця')
            back = types.KeyboardButton('🔙Назад')
            keyboard.add(back)
            keyboard.add(item1, item2, item3, item4, item5)

            bot.send_message(message.chat.id, "Оберіть день, у який ви внесете корекцію", reply_markup=keyboard)
            bot.register_next_step_handler(message, Menu_dess.red_rozklad_2, text)

        else:
            bot.send_message(message.chat.id, "Такого варіанту немає")
            bot.register_next_step_handler(message, Menu_dess.red_rozklad, text)
    @staticmethod
    def red_rozklad_2(message, text):
        den = message.text
        if den == "🔙Назад":
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
            markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
            markup.add(types.KeyboardButton("Додати диспетчера"), types.KeyboardButton("Додати навчальні плани"))
            markup.add(types.KeyboardButton("Кількість користувачів"), types.KeyboardButton("Редагувати студента"))
            bot.send_message(message.chat.id, "Меню диспетчера", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        elif den not in ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'П\'ятниця']:
            bot.send_message(message.chat.id, "Ви ввели не правильний день, будь ласка виберіть день із наявних кнопок: ")
            bot.register_next_step_handler(message, Menu_dess.red_rozklad_2, text)
        else:
            days = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'П\'ятниця']
            den_123 = days.index(den)

            # Встановлюємо з'єднання з базою даних
            conn = sqlite3.connect('users.db')
            cursor = conn.cursor()
            text = text.replace("-","_")
            cursor.execute(f'SELECT * FROM rosklad_{text}')
            # Отримуємо результат запиту (перший рядок таблиці)
            first_row = cursor.fetchall()[den_123]

            mess = ''
            for j, item in enumerate(first_row):
                if item is not None:
                    mess += f"{j + 1}. {item}\n"

            bot.send_message(message.chat.id, f'\n<code>{mess}</code>', parse_mode=ParseMode.HTML)
            user_grypa = text
            bot.register_next_step_handler(message, Menu_dess.red_rozklad_3, user_grypa ,den_123)
            # Закриваємо з'єднання з базою даних
            cursor.close()
            conn.close()
            bot.send_message(message.chat.id,f"Ви вибрали {den}, будь ласка надішліть відредагований розкла.\nЗА ТАКИМ ЗРАЗКОМ!!!")
            # Продовження допиши як правильно вставляти
    @staticmethod
    def red_rozklad_3(message, user_grypa,den_123):
        text = message.text  # Отримуємо текст повідомлення
        if text == '🔙Назад':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            markup.add(types.KeyboardButton('Перегляд оцінок'), types.KeyboardButton('Редагувати розклад'))
            markup.add(types.KeyboardButton('Обновлення пошт'), types.KeyboardButton('Редагування списку груп'))
            markup.add(types.KeyboardButton("Додати диспетчера"), types.KeyboardButton("Додати навчальні плани"))
            markup.add(types.KeyboardButton("Кількість користувачів"), types.KeyboardButton("Редагувати студента"))
            bot.send_message(message.chat.id, "Меню диспетчера", reply_markup=markup)
            bot.register_next_step_handler(message, Menu_dess.menu_desp)
        else:
            try:
                # Розбиваємо текст на окремі рядки
                lines = text.split('\n')
                result = [None] * 5  # Результат (кортеж) з початковим значенням None
                for line in lines:
                    # Розбиваємо рядок на індекс і значення
                    index, value = line.split('. ', 1)
                    # Перевіряємо, чи індекс є числом від 1 до 5
                    if index.isdigit() and 1 <= int(index) <= 5:
                        result[int(index) - 1] = value
                # Виводимо результат (кортеж)
                result = tuple(result)

                conn = sqlite3.connect('users.db')
                cursor = conn.cursor()
                cursor.execute(f'SELECT * FROM rosklad_{user_grypa}')
                rows = cursor.fetchall()
                rows[den_123] = result
                cursor.execute(f'DELETE FROM rosklad_{user_grypa}')
                cursor.executemany(f'INSERT INTO rosklad_{user_grypa} VALUES (?,?,?,?,?)', rows)
                conn.commit()
                cursor.close()
                conn.close()
                bot.send_message(message.chat.id, "Розклад успішно оновленно")
                connect = sqlite3.connect('users.db')
                cursor = connect.cursor()
                cursor.execute("SELECT Групи FROM Групи")
                rows = cursor.fetchall()
                gryps = [row[0] for row in rows]
                columns = 3
                gryps_per_column = (len(gryps) + columns - 1) // columns
                gryps_divided = [gryps[i:i + gryps_per_column] for i in range(0, len(gryps), gryps_per_column)]

                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                markup.add(types.KeyboardButton('🔙Назад'))
                # Додавання груп до розмітки по 3 в кожному рядку
                for gryp_column in gryps_divided:
                    markup.add(*gryp_column)

                connect.close()
                bot.send_message(message.chat.id, 'Оберіть групу в якій хочете редагувати розклад:',reply_markup=markup)
                bot.register_next_step_handler(message, Menu_dess.red_rozklad, gryps)
            except Exception as e:
                bot.send_message(message.chat.id,'Ви ввели не по зразку, будь ласка надішліть мені ще раз відредагований розклад за зразком')
                bot.send_message(message.chat.id, '<code>1. (Назва пари)\n2. (Назва пари)\n3. (Назва пари)\n4. (Назва пари)\n5. (Назва пари)</code>',parse_mode=ParseMode.HTML)
                bot.register_next_step_handler(message, Menu_dess.red_rozklad_3, user_grypa ,den_123)



@bot.message_handler(content_types=['text'])
def bot_message1(message,text):
    if text == 'я староста' or text == "Я староста":
        bot.send_message(message.chat.id, "Ти піздюк, а не староста😏")
        bot.forward_message(chat_id=CHAT_ID, from_chat_id=message.chat.id, message_id=message.message_id)

    elif "путін" in text.lower():
        bot.send_message(message.chat.id, "Хуйло🤝")
        bot.forward_message(chat_id=CHAT_ID, from_chat_id=message.chat.id, message_id=message.message_id)


    elif "путінхуйло" in text or "путін хуйло" in text:
        bot.send_message(message.chat.id, "Згідний🤝")
        bot.forward_message(chat_id=CHAT_ID, from_chat_id=message.chat.id, message_id=message.message_id)


    elif "слава україні" in text.lower():
        bot.send_message(message.chat.id, "Героям слава!")
        bot.forward_message(chat_id=CHAT_ID, from_chat_id=message.chat.id, message_id=message.message_id)

    elif "прігожин" in text.lower():
        bot.send_message(message.chat.id, "Долітався")
        bot.forward_message(chat_id=CHAT_ID, from_chat_id=message.chat.id, message_id=message.message_id)


    else:
        bot.send_message(message.chat.id, f"Такого варіанту немає")
        bot.forward_message(chat_id=CHAT_ID, from_chat_id=message.chat.id, message_id=message.message_id)


def chet_teacer(message):
    user_id = message.chat.id
    user_id = str(user_id)
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f"SELECT id FROM Викладачі")
    results = cursor.fetchall()
    results = [item for tpl in results for item in tpl]

    if user_id not in results:
        cursor.execute(f"SELECT email, grypa FROM login_id WHERE id = {user_id}")
        data = cursor.fetchall()


        bot.send_message(message.chat.id, 'Вибачте, але вас ще немає в базі, будь ласка пройдіть реєстрацію предметів і назв предметів(повну назву навчальної дисципліни) і через дефіс Групу в якій викладаєте його за таким зразком ‼️‼️\nЗразок:\nЕлементарна математика - СОМІ-23\nМатематичний аналіз - СОФА-35\nПредмет - Група\nПредмет - Група')
        creat_teacer(message,user_id,data)
        conn.close()
    else:
        conn.close()
        menu_vikladacham(message)
def creat_teacer(message,user_id,data):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f'CREATE TABLE IF NOT EXISTS "{user_id}_tea" (предмет TEXT, група TEXT)')
    for row in data:
        email, kaf = row
        cursor.execute('INSERT INTO "Викладачі" (id, email, Кафедра) VALUES (?,?,?);',(user_id, email,kaf))
    conn.commit()
    conn.close()
    bot.register_next_step_handler(message, teacher_pred,user_id)
def teacher_pred(message,user_id):
    pred = message.text
    lines = pred.split('\n')
    lines = [i.split(" - ") for i in lines]

    list = []
    for i in lines:
        if len(i) != 2:
            list.append(i)
            lines.remove(i)

    message_text = "\n\n".join(str(el) for el in list)
    message_text2 = "\n".join([f'{item[0]} - {item[1]}' for item in lines])
    if len(list) > 0:
        if len(lines) > 0:
            bot.send_message(message.chat.id,f"{message_text}\n тут виведені рядок, які Ви ввели не правильно\n\n{message_text2}\nА ось тут правильні\nВведіть будь ласка ще раз свої предмети і групи за зразком")
            bot.register_next_step_handler(message, teacher_pred, user_id)
        elif len(lines) == 0:
            bot.send_message(message.chat.id,f"{message_text}\n тут виведені рядок, які Ви ввели не правильно.Введіть будь ласка ще раз свої предмети і групи за зразком")
            bot.register_next_step_handler(message, teacher_pred, user_id)
    else:
        conn = sqlite3.connect("users.db")
        cursor = conn.cursor()
        cursor.execute("SELECT ПредметиФізмату FROM ПредметиФізмату")
        result = cursor.fetchall()
        result = [item.lower() for tpl in result for item in tpl]
        connect = sqlite3.connect('users.db')
        cursor = connect.cursor()
        cursor.execute("SELECT Групи FROM Групи")
        rows = cursor.fetchall()
        gryp = [row[0] for row in rows]


        fakepred = ''
        lines_list = []
        gryp = [i.upper() for i in gryp]

        for i in lines:
            subject, grypa = i
            if subject.lower() not in result or grypa.upper() not in gryp:
                fakepred += subject +' - ' + grypa + '\n'
            else:
                lines_list.append(i)
        if len(fakepred)>0:
            bot.send_message(message.chat.id, f'{fakepred} ось ці рядки введене не правильно або неправильно вказаний предмет або не вірна назва групи\nВведіть ще раз предмет і групу за зразком')
            bot.register_next_step_handler(message, teacher_pred, user_id)
        else:
            teacher_pred2(message, user_id, lines_list)
def teacher_pred2(message, user_id, lines_list):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    for i in lines_list:
        subject, grypa = i
        subject = subject.lower()
        grypa = grypa.upper()
        esc = ' _<>,.?!@#$%^&*()+=`"\''
        for char in esc:
            grypa = grypa.replace(char, '-')
        cursor.execute(f'INSERT INTO "{user_id}_tea" (предмет, група) VALUES (?,?);', (subject, grypa,))
    conn.commit()
    conn.close()
    menu_vikladacham(message)
def ogoloshennya_grypa(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item_back = types.KeyboardButton('🔙Назад')
    markup.add(item_back)
    bot.send_message(chat_id=message.chat.id, text='Введіть повідомлення для розсилки своїй групі:',reply_markup= markup)
    bot.register_next_step_handler(message, ogoloshennya_grypa2)
def ogoloshennya_grypa2(message):
    news = message.text
    if news == '🔙Назад':
        message_handler_start(message)

    else:
        user_id = message.chat.id
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT grypa FROM login_id WHERE id = {user_id}")
        user_grypa = cursor.fetchone()[0]
        cursor.execute(f"SELECT id FROM login_id WHERE grypa = '{user_grypa}'")
        all = cursor.fetchall()
        all = [item for tpl in all for item in tpl]
        blocked_users = []

        if message.content_type == 'text':
            for user in all:
                try:
                    bot.forward_message(chat_id=user, from_chat_id=message.chat.id, message_id=message.message_id)
                except telebot.apihelper.ApiTelegramException as e:
                    if e.result.status_code == 403:
                        blocked_users.append(user)
            # Закріплюємо повідомлення в чаті
            try:
                pinned_message = bot.send_message(chat_id=message.chat.id, text=message.text)
                bot.pin_chat_message(chat_id=message.chat.id, message_id=pinned_message.message_id)
            except telebot.apihelper.ApiTelegramException:
                pass
            message_handler_start(message)
        else:
            for user in all:
                try:
                    bot.forward_message(chat_id=user, from_chat_id=message.chat.id, message_id=message.message_id)
                except telebot.apihelper.ApiTelegramException as e:
                    if e.result.status_code == 403:
                        blocked_users.append(user)
            # Закріплюємо повідомлення в чаті
            try:
                pinned_message = bot.send_message(chat_id=message.chat.id, text=news)
                bot.pin_chat_message(chat_id=message.chat.id, message_id=pinned_message.message_id)
            except telebot.apihelper.ApiTelegramException:
                pass
            message_handler_start(message)

        if blocked_users:
            blocked = []
            for i in blocked_users:
                cursor.execute(f"SELECT email, first_last FROM login_id WHERE id = {i}")
                result = cursor.fetchone()
                blocked.append(result)
            blocked_users_text = ''
            for i in blocked:
                text = ': '.join(i)
                blocked_users_text += text + "\n"
            # Надсилання повідомлення про заблокованих користувачів адміністратору бота
            bot.send_message(message.chat.id, f"Твої одногрупники, які заблокували бота і не отримали оголошення\n{blocked_users_text}")
            message_handler_start(message)

def pereglad_ocinok(message, db_filename,first_last):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute('SELECT Предмети FROM Предмети')
    result = cursor.fetchall()

    keyboard = []
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('🔙Назад'))
    for row in result:
        subject = row[0]
        keyboard.append(subject)
        markup.add(types.KeyboardButton(subject))



    bot.send_message(message.chat.id, "Оберіть предмет з якого хочете переглянути свої оцінки", reply_markup=markup)
    bot.register_next_step_handler(message,pereglad_ocinok_2,db_filename,first_last, keyboard)
def pereglad_ocinok_2(message, db_filename,first_last, keyboard):
    subject = message.text
    if subject == "🔙Назад":
        message_handler_start(message)

    elif subject in keyboard:
        subject = subject.replace(" ", "_")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        table_name = f'{subject}_Студенти'
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if
                        column[1] != 'Студенти' and column[1] != 'модуль_1' and column[1] != 'модуль_2' and column[
                            1] != 'Н' and column[1] != 'Індз' and column[1] != 'Загальна_кількість_балів']

        query = f"SELECT модуль_1, модуль_2, Індз, {column_names[0]}, Загальна_кількість_балів FROM {subject}_Студенти WHERE Студенти = ?"
        cursor.execute(query, (first_last,))
        row = cursor.fetchone()

        if row is not None:
            module_1 = row[0] if row[0] is not None else '0'
            module_2 = row[1] if row[1] is not None else '0'
            indz = row[2] if row[2] is not None else '0'
            exam = row[3] if row[3] is not None else '0'
            total_score = row[4] if row[4] is not None else '0'

            result_string = (
                f"Модуль 1: {module_1}\n"
                f"Модуль 2: {module_2}\n"
                f"ІНДЗ(якщо цього семестру є): {indz}\n"
                f"{column_names[0]}: {exam}\n"
                f"Загальна кількість балів: {total_score}"
            )
            bot.send_message(message.chat.id,result_string)
            bot.register_next_step_handler(message, pereglad_ocinok_2, db_filename, first_last, keyboard)
        else:
            bot.send_message(message.chat.id, "Вас не знайдено, можливо ви зареєстровані під неправильним ПІБ")
            message_handler_start(message)



    else:
        bot.send_message(message.chat.id, "Такого варіанту відповіді немає")
        bot.register_next_step_handler(message, pereglad_ocinok_2, db_filename, first_last, keyboard)

def jurnal1(message):
    user_id = message.chat.id
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f"SELECT grypa FROM login_id WHERE id = {user_id}")
    user_grypa = cursor.fetchone()[0]
    user_grypa = user_grypa.upper().replace('_', '-')

    db_filename = user_grypa + '.db'
    if os.path.exists(db_filename):                 #1_1234
        bot.send_message(message.chat.id,f"Журнал групи {user_grypa} вже створений.")
        jurnal1_1(message)

    else:              #2_1234
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('Створити'))
        markup.add(types.KeyboardButton('Відмінити'))
        bot.send_message(message.chat.id, f'Ви впевнені, що хочете зараз створити журнал?\n‼️УВАГА‼️\nЯкщо нажмете кнопку "Створити" ви не зможете повернутись в меню, допоки не добавите список студентів, та предмети, які вам читають.',
                         reply_markup=markup)
        bot.register_next_step_handler(message, jurnal2, user_grypa)
def jurnal2(message,user_grypa):
    text = message.text
    if text == 'Створити':
        bot.send_message(message.chat.id, f"Давайте я допоможу вам створити журнал для вашої групи {user_grypa}",reply_markup=telebot.types.ReplyKeyboardRemove())
        Create_jurnal.jurnal2_1(message, user_grypa)
    elif text == 'Відмінити':
        menu_starostam(message)
    else:
        bot.send_message(message.chat.id, f"Такого варіанту відповіді немає")
        bot.register_next_step_handler(message, jurnal2, user_grypa)
def jurnal1_1(message):
    user_id = message.chat.id
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f"SELECT grypa FROM login_id WHERE id = {user_id}")
    user_grypa = cursor.fetchone()[0]
    user_grypa = user_grypa.upper().replace('_', '-')

    db_filename = user_grypa + '.db'
    conn.close()
    jurnal1_2_1interval(message, db_filename, user_grypa)
def jurnal1_2_1interval(message, db_filename, user_grypa):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute("SELECT Лекційний_Тиждень FROM Лекційний_Тиждень")
    close = cursor.fetchone()[0]
    if close == "Закритий":
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        markup.add(types.KeyboardButton('Отримати таблицю'))
        markup.add(types.KeyboardButton('Додати предмет📚'), types.KeyboardButton('Робота з предметами📚'))
        markup.add(types.KeyboardButton('Додати Студента🎓'), types.KeyboardButton('Перегляд списку групи🎓'))
        bot.send_message(message.chat.id, "Оберіть функцію, з меню кнопок", reply_markup=markup)
        conn.commit()
        conn.close()
        bot.register_next_step_handler(message, jurnal1_2_2interval, db_filename, user_grypa)

    else:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        markup.add(types.KeyboardButton('Лекційний тиждень'))
        markup.add(types.KeyboardButton('Отримати таблицю'))
        markup.add(types.KeyboardButton('Додати предмет📚'),types.KeyboardButton('Робота з предметами📚'))
        markup.add(types.KeyboardButton('Додати Студента🎓'),types.KeyboardButton('Перегляд списку групи🎓'))

        bot.send_message(message.chat.id, "Оберіть функцію, з меню кнопок",reply_markup=markup)
        conn.commit()
        conn.close()
        bot.register_next_step_handler(message, jurnal1_2_2interval, db_filename,user_grypa)
def jurnal1_2_2interval(message, db_filename, user_grypa):
    text = message.text
    if text == '🔙Назад':
        menu_starostam(message)

    elif text == 'Отримати таблицю':
        conn = sqlite3.connect(f'{db_filename}')
        cursor = conn.cursor()
        cursor.execute('SELECT Предмети FROM Предмети')
        result = cursor.fetchall()

        keyboard = []
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        for row in result:
            subject = row[0]
            keyboard.append(subject)
            markup.add(types.KeyboardButton(subject))

        conn.close()
        bot.send_message(message.chat.id, "Оберіть предмет з якого ви хочете отримати таблицю", reply_markup=markup)
        bot.register_next_step_handler(message, Create.create_table, db_filename, user_grypa)

    elif text == 'Лекційний тиждень':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Переглянути розклад')
        item2 = types.KeyboardButton('Виставити Н')
        item3 = types.KeyboardButton('Закрити Л. тиждень')
        markup.add(types.KeyboardButton('🔙Назад'))
        markup.add(item1)
        markup.add(item2)
        markup.add(item3)
        bot.send_message(message.chat.id, "Оберіть, що хочете зробити із лекційним тижнем ", reply_markup=markup)
        bot.register_next_step_handler(message, less, db_filename, user_grypa)
    elif text == 'Додати предмет📚':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Так👌')
        item2 = types.KeyboardButton('Ні👎🏿')
        markup.add(item1)
        markup.add(item2)
        bot.send_message(message.chat.id,"Оуу окей ви вирішили добавити предмет у свій журнал ви впевнені що хочете це зробити?",reply_markup=markup)
        bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_1, db_filename, user_grypa)
    elif text == 'Додати Студента🎓':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Так👌')
        item2 = types.KeyboardButton('Ні👎🏿')
        markup.add(item1)
        markup.add(item2)
        bot.send_message(message.chat.id,"Оуу окей ви вирішили добавити нового студента у свій журнал ви впевнені, що хочете це зробити?",reply_markup=markup)
        bot.register_next_step_handler(message, jurnal1_3_student_add, db_filename)
    elif text == 'Перегляд списку групи🎓':
        conn = sqlite3.connect(f'{db_filename}')
        cursor = conn.cursor()
        cursor.execute('SELECT Студенти FROM STUDENTY')
        result = cursor.fetchall()
        result_string = '\n'.join(item[0] for item in result)
        bot.send_message(message.chat.id,f"Ось список вашої групи\n\n{result_string}")
        bot.register_next_step_handler(message, jurnal1_2_2interval, db_filename, user_grypa)
    elif text == 'Робота з предметами📚':
        jurnal1_2(message, db_filename, user_grypa)
def less(message,db_filename, user_grypa):
    text = message.text
    if text == '🔙Назад':
        jurnal1_2_1interval(message, db_filename, user_grypa)
    elif text == 'Переглянути розклад':
        edit_less(message,db_filename, user_grypa)

    elif text == 'Виставити Н':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Понеділок')
        item2 = types.KeyboardButton('Вівторок')
        item3 = types.KeyboardButton('Середа')
        item4 = types.KeyboardButton('Четвер')
        item5 = types.KeyboardButton('П\'ятниця')
        item6 = types.KeyboardButton('Субота')
        markup.add(types.KeyboardButton('🔙Назад'))
        markup.add(item1, item2, item3)
        markup.add(item4, item5, item6)
        bot.send_message(message.chat.id, "Оберіть день в якому хочете виставити Н",reply_markup=markup)
        bot.register_next_step_handler(message, grate_less, db_filename, user_grypa)

    elif text == 'Закрити Л. тиждень':
        close_less(message,db_filename, user_grypa)
def grate_less(message,db_filename, user_grypa):
    day = message.text
    if day == '🔙Назад':
        jurnal1_1(message)
    elif day in ['Понеділок','Вівторок','Середа','Четвер','П\'ятниця','Субота']:
        day = day.replace("\'", "")
        table_name = f"лекційний_{day}"
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if column[1] != 'Студенти']
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        for column_name in column_names:
            markup.add(column_name)
        bot.send_message(message.chat.id, "Оберіть предмет в якому хочете виставити Н", reply_markup=markup)
        conn.close()
        bot.register_next_step_handler(message, grate_less_2, db_filename, user_grypa, column_names,table_name)
    else:
        bot.send_message(message.chat.id, "Такого варіанту немає\nОберіть предмет в якому хочете виставити Н")
        bot.register_next_step_handler(message, grate_less, db_filename, user_grypa)
def grate_less_2(message,db_filename, user_grypa, column_names,table_name):
    pred = message.text
    if pred in column_names:
        bot.send_message(message.chat.id,
                         "Оцінки не ставте а ставте тільки Н якщо студент був присутній то пропуск")
        bot.send_message(message.chat.id,
                         "ПІБ(Одногрупника) - Н\nПІБ(Одногрупника) - \nПІБ(Одногрупника) - Н")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()

        # Витягуємо всі значення зі стовпця "Студенти" таблиці "STUDENTY"
        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()

        # Формуємо повідомлення зі списком студентів
        students_list = "\n".join([student[0] + ' - 0' for student in students])

        # Надсилаємо повідомлення зі списком студентів у бота
        bot.send_message(message.chat.id, f"<code>{students_list}</code>", parse_mode=ParseMode.HTML)
        conn.close()
        bot.register_next_step_handler(message, grate_less_3, db_filename, user_grypa, pred,table_name)
    else:
        bot.send_message(message.chat.id,
                         "Такого варіанту немає")
        bot.register_next_step_handler(message, grate_less_2, db_filename, user_grypa, column_names,table_name)
def grate_less_3(message,db_filename, user_grypa, pred,table_name):
    text = message.text
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    text = text.split("\n")
    split = []


    for i in text:

        student_data = i.split(" - ")
        if len(student_data) == 2:
            split.append(student_data)

    spec = "_<>,.?!@#$%^&*()+=\""
    for i in split:
        for k,w in enumerate(i):
            if w in spec or w == "" or w == " ":
                i[k] = None

    splitnot =[]
    split2 = []
    for i in split:
        for k,w in enumerate(i):
            if w.isdigit():
                splitnot.append(i[k])

            else:
                split2.append(i[k])

    if len(splitnot) > 0:
        list = tuple(splitnot)
        message_text = "\n\n".join(list)
        bot.send_message(message.chat.id,f"{message_text} Ось ці рядки я не зміг розпізнати, будь ласка надішліть ще раз і правильно за таким зразком\n\nПІБ - Н\nПІБ - ")
        bot.register_next_step_handler(message, grate_less_3, db_filename, user_grypa)

    else:

        for row in split:
            name, grade = row
            cursor.execute(f"UPDATE {table_name} SET [{pred}] = ? WHERE Студенти = ?",(grade, name))
        conn.commit()
        conn.close()
def edit_less(message,db_filename, user_grypa):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton('Понеділок')
    item2 = types.KeyboardButton('Вівторок')
    item3 = types.KeyboardButton('Середа')
    item4 = types.KeyboardButton('Четвер')
    item5 = types.KeyboardButton('П\'ятниця')
    item6 = types.KeyboardButton('Субота')
    markup.add(types.KeyboardButton('🔙Назад'))
    markup.add(item1,item2,item3)
    markup.add(item4,item5,item6)
    bot.send_message(message.chat.id, "Оберіть день в якому хочете відредагувати розклад", reply_markup=markup)
    bot.register_next_step_handler(message, edit_less_2, db_filename, user_grypa)
def edit_less_2(message,db_filename, user_grypa):
    text = message.text
    if text == 'П\'ятниця':
        text1 = 'П\'ятниця'
        text = text.replace("`", '')
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info(лекційний_{text})")
        columns = cursor.fetchall()
        columns_names = [item[1] for item in columns]
        columns_names.remove(columns_names[0])
        formatted_strings = []
        for i, subject in enumerate(columns_names):
            formatted_strings.append(f"{i + 1}) {subject}")
        formatted_output = "\n".join(formatted_strings)
        bot.send_message(message.chat.id, f"Ось розклад на {text1}\n{formatted_output}",
                         reply_markup=telebot.types.ReplyKeyboardRemove())
        time.sleep(1)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Редагувати')
        item2 = types.KeyboardButton('🔙Назад')
        markup.add(item2)
        markup.add(item1)
        bot.send_message(message.chat.id,
                         f'Можливо вам потрібно відредагувати лекційний розклад на {text}, якщо так то надішліть мені "Редагувати", якщо ні то можете повернутися кнопкою "🔙Назад"',
                         reply_markup=markup)
        bot.register_next_step_handler(message, edit_less_3, db_filename, user_grypa, text)
    elif text in ['Понеділок','Вівторок','Середа','Четвер','Субота']:
        text1 = text
        text = text.replace("`", '')
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info(лекційний_{text})")
        columns = cursor.fetchall()
        columns_names = [item[1] for item in columns]
        columns_names.remove(columns_names[0])
        formatted_strings = []
        for i, subject in enumerate(columns_names):
            formatted_strings.append(f"{i + 1}) {subject}")
        formatted_output = "\n".join(formatted_strings)
        bot.send_message(message.chat.id, f"Ось розклад на {text1}\n{formatted_output}",
                         reply_markup=telebot.types.ReplyKeyboardRemove())
        time.sleep(1)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Редагувати')
        item2 = types.KeyboardButton('🔙Назад')
        markup.add(item2)
        markup.add(item1)
        bot.send_message(message.chat.id,
                         f'Можливо вам потрібно відредагувати лекційний розклад на {text}, якщо так то надішліть мені "Редагувати", якщо ні то можете повернутися кнопкою "🔙Назад"',
                         reply_markup=markup)
        bot.register_next_step_handler(message, edit_less_3, db_filename, user_grypa, text)

    elif text == '🔙Назад':
        jurnal1(message)
    else:
        bot.send_message(message.chat.id, f"Такого варіанту немає, оберіть ще раз")
        bot.register_next_step_handler(message, edit_less_2, db_filename, user_grypa)
def edit_less_3(message,db_filename, user_grypa,text):
    txt = message.text
    day = text
    if txt == 'Редагувати':
        bot.send_message(message.chat.id,
                         f"Що ж ви обрали редагування {text} у лекційному тижні надішліть будь ласка мені за таким зразком через абзац свої предмети і викладачів це необхідно для майбутньої документаціх\n‼️Якщо у вас немає наприклад 2 пари замість назви пари напишіть (пара відсутня)‼️")
        bot.send_message(message.chat.id,
                         "Предмет - ПІБ Викладача\nПредмет - ПІБ Викладача\nПредмет - ПІБ Викладача\nПредмет - ПІБ Викладача\n",reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message,edit_less_4, db_filename, user_grypa,day)
    elif txt == '🔙Назад':
        jurnal1_1(message)
    else:
        bot.send_message(message.chat.id, f"Такого варіанту немає")
        bot.register_next_step_handler(message, edit_less_3, db_filename, user_grypa, text)
def remove_numbering(text_list):
    return [line.split(". ")[1] if ". " in line else line for line in text_list]
def edit_less_4(message,db_filename, user_grypa, day):
    text = message.text
    text = text.split("\n")
    text = remove_numbering(text)
    if len(text) != 4:
        bot.send_message(message.chat.id, f"Має бути введено рівно 4 предмети. Введіть ще раз.",
                         reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, edit_less_4, db_filename, user_grypa, day)
    else:
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info(лекційний_{day})")
        columns = cursor.fetchall()
        columns_names = [item[1] for item in columns]
        columns_names.remove(columns_names[0])
        t=0
        for i in columns_names:
            cursor.execute(f"ALTER TABLE лекційний_{day} RENAME COLUMN [{i}] TO [{text[t]}]")
            t+=1

        # Відправлення підтвердження
        bot.send_message(message.chat.id, "Розклад на день оновлено.", reply_markup=telebot.types.ReplyKeyboardRemove())
        jurnal1_2_1interval(message, db_filename, user_grypa)
def close_less(message,db_filename, user_grypa):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton('Так👌')
    item2 = types.KeyboardButton('Ні👎🏿')
    markup.add(item1)
    markup.add(item2)
    bot.send_message(message.chat.id,"Ви точно хочете закрити лекційний тиждень?",reply_markup=markup)
    bot.register_next_step_handler(message, close_less_2, db_filename, user_grypa)
def close_less_2(message,db_filename, user_grypa):
    text = message.text
    if text == 'Так👌':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        subject = "Закритий"
        cursor.execute(f'UPDATE Лекційний_Тиждень SET Лекційний_Тиждень = ?',(subject,))
        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, f"Лекційний тиждень закритий, якщо вам необхідно буде відкрити його зверніться в /support, і подайте заявку на відкриття лекційного тижня")
        jurnal1_2_1interval(message, db_filename, user_grypa)
    if text == 'Ні👎🏿':
        bot.send_message(message.chat.id,f"Повертаюся в меню робота з журналом...")
        jurnal1_2_1interval(message, db_filename, user_grypa)
def jurnal1_2(message,db_filename,user_grypa):
    conn = sqlite3.connect(f'{db_filename}')
    cursor = conn.cursor()
    cursor.execute('SELECT Предмети FROM Предмети')
    result = cursor.fetchall()

    keyboard = []
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('🔙Назад'))
    for row in result:
        subject = row[0]
        keyboard.append(subject)
        markup.add(types.KeyboardButton(subject))

    conn.close()
    bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
    bot.register_next_step_handler(message, jurnal1_3, db_filename,user_grypa,keyboard)
def jurnal1_3(message,db_filename,user_grypa,keyboard):
    subject = message.text
    if subject == '🔙Назад':
        jurnal1_2_1interval(message, db_filename, user_grypa)

    elif subject.startswith('/'):
        bot.send_message(message.chat.id, "Ви ввели команду будь ласка оберіть предмет:")
        jurnal1_2(message, db_filename, user_grypa)




    elif subject in keyboard:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Модуль 1')
        item2 = types.KeyboardButton('Модуль 2')
        item3 = types.KeyboardButton('ІНДЗ')
        back = types.KeyboardButton('🔙Назад')
        markup.add(back)
        markup.add(item1)
        markup.add(item2)
        markup.add(item3)
        bot.send_message(message.chat.id, "Оберіть, розділ:",reply_markup=markup)
        bot.register_next_step_handler(message, jurnal1_4, db_filename, user_grypa, subject)

    else:
        bot.send_message(message.chat.id,"Такого варіанту відповіді немає оберіть ще разок")
        jurnal1_1(message)
def jurnal1_3_student_add(message,db_filename):
    text = message.text
    if text == 'Так👌':
        bot.send_message(message.chat.id,"Зараз ви можете добавити лише одного студента у свій журнал.\nНадішліть мені його Прізвище Ім\'я По-батькові",reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, jurnal1_3_student_add_2, db_filename)

    elif text == 'Ні👎🏿':
        bot.send_message(message.chat.id, "Повертаюся назад...")
        time.sleep(2)
        jurnal1_1(message)
    else:
        bot.send_message(message.chat.id, "Такого варіанту відповіді немає оберіть ще разок")
        bot.register_next_step_handler(message, jurnal1_3_student_add, db_filename)
def jurnal1_3_student_add_2(message,db_filename):
    text = message.text
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute("SELECT Студенти FROM STUDENTY")
    spisok = cursor.fetchall()
    spisok = [item.lower() for tpl in spisok for item in tpl]
    conn.close()
    if has_special_characters(text):
        bot.send_message(message.chat.id, "Неправильний формат рядка навіщо ти надішслав спец символи?\nНадішли ще раз ПІБ нового студента")
        bot.register_next_step_handler(message, jurnal1_3_student_add_2, db_filename)

    elif text.lower() in spisok:
        id = message.chat.id
        user_name = message.chat.username
        bot.send_message(message.chat.id,"Такий студент вже є у вашому журналі не додавайте ліпових студентів")
        bot.send_message(CHAT_ID, f"Староста під айді: {id}\nНіком: @{user_name} \nХотіла додати ліпового студента")
        jurnal1_1(message)

    else:
        jurnal1_3_student_add_3(message,db_filename, text)
def has_special_characters(text):
    special_characters = ".,/\\|?!@$%^:;&*()_+=[]{}\n\"AaBbCcDdEeFfGgHhIiJjKkLlMmNnOoPpQqRrSsTtUuVvWwXxYyZz"
    for char in text:
        if char in special_characters:
            return True
    return False
def jurnal1_3_student_add_3(message,db_filename, text):
    words = text.split()

    # Змінити регістр першої літери кожного слова (велика літера)
    formatted_words = [word.capitalize() for word in words]

    # Приєднати слова знову, вставивши пробіл між ними
    text = " ".join(formatted_words)
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute("SELECT Предмети FROM Предмети")
    res = cursor.fetchall()
    res = [item.lower() for tpl in res for item in tpl]
    for i in res:
        i = i.replace(" ", "_")
        cursor.execute(f"PRAGMA table_info({i}_1)")
        columns = cursor.fetchall()

        column_names = ['[' + column[1] + "]" for column in columns]

        st = ", ".join(column_names)
        j = []

        for n in range(len(column_names)-1):
            j.append("NULL")
        j_st = ", ".join(j)
        cursor.execute(f'INSERT INTO "{i}_1" ({st}) VALUES ("{text}", {j_st})')




        cursor.execute(f"PRAGMA table_info({i}_2)")
        columns = cursor.fetchall()

        column_names = ['[' + column[1] + "]" for column in columns]

        st = ", ".join(column_names)
        j=[]
        for n in range(len(column_names)-1):
            j.append("NULL")
        j_st = ", ".join(j)
        cursor.execute(f'INSERT INTO "{i}_2" ({st}) VALUES ("{text}", {j_st})')

        cursor.execute(f"PRAGMA table_info([{i}_3])")
        columns_info = cursor.fetchall()
        columns_names = [column_info[1] for column_info in columns_info]
        exam_type = columns_names[2]

        cursor.execute(f'INSERT INTO "{i}_3" (Студенти, Індз, [{exam_type}],  модуль_1, модуль_2, Загальна_кількість_балів) VALUES ("{text}", NULL, NULL, NULL, NULL, NULL)')



        cursor.execute(f"PRAGMA table_info({i}_Індивідуальні_години)")
        columns = cursor.fetchall()

        column_names = ['[' + column[1] + "]" for column in columns]
        st = ", ".join(column_names)
        j = []
        for n in range(len(column_names) - 1):
            j.append("NULL")
        j_st = ", ".join(j)
        if len(column_names)>1:
            cursor.execute(f'INSERT INTO "{i}_Індивідуальні_години" ({st}) VALUES ("{text}", {j_st})')

        else:
            cursor.execute(f'INSERT INTO "{i}_Індивідуальні_години" (Студенти) VALUES ("{text}")')






        cursor.execute(f'INSERT INTO "{i}_Студенти" ("Студенти", модуль_1, модуль_2, [Індз],[{exam_type}], Загальна_кількість_балів) VALUES ("{text}", NULL, NULL, NULL, NULL, NULL)')

    cursor.execute(f'INSERT INTO "STUDENTY" (Студенти) VALUES ("{text}")')
    bot.send_message(message.chat.id,"Студента успішно додано надіюсь ви ввели правильно ПІБ студента)")
    i = db_filename.split("-")
    user_grypa = i[0]
    jurnal1_2_1interval(message, db_filename, user_grypa)
    conn.commit()
    conn.close()
def jurnal1_3_dodavanya_predmety_1(message, db_filename, user_grypa):
    text = message.text
    if text == 'Так👌':
        bot.send_message(message.chat.id,"Максимум зараз ви можете добавити лише ОДИН предмет за ОДИН раз\nВведіть назву навчальної дисципліни за таким зразком\nБудь уважні та робіть пробіли між словами на символами як у зразку\nЗРАЗОК:\nМатематичний аналіз - 4, залік\n\nФорма запису:\nПредмет - ЄКТС, Тип підсумкового контролю", reply_markup=telebot.types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)

    elif text == 'Ні👎🏿':
        bot.send_message(message.chat.id, "Повертаюся назад...")
        time.sleep(2)
        jurnal1_1(message)
    else:
        bot.send_message(message.chat.id, "Такого варіанту відповіді немає оберіть ще разок")
        jurnal1_1(message)
def jurnal1_3_dodavanya_predmety_2(message, db_filename):
    pred = message.text
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    cursor.execute("SELECT ПредметиФізмату FROM ПредметиФізмату")
    result = cursor.fetchall()
    result = [item.lower() for tpl in result for item in tpl]

    lines = pred.split('\n')
    if len(lines) > 1:
        bot.send_message(message.chat.id, "Ви ввели забагато предметів зараз ви можете додати лише один предмет")
        bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)
    else:
        elements = pred.split(' - ')

        if len(elements) != 2:
            bot.send_message(message.chat.id, f"Не правильний формат рядка\n{pred}\nВведіть за зразком Предмет - ЄКТС, Тип підсумкового контролю 1")
            bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)
        else:
            infopred = [i for el in elements for i in el.split(",")]
            infopred = [item.strip().lower() for item in infopred]

            if len(infopred) !=3:
                bot.send_message(message.chat.id, f"Не правильний формат рядка\n{pred}\nВведіть за зразком Предмет - ЄКТС, Тип підсумкового контролю 2")
                bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)
            else:

                if infopred[0].lower() in result:
                    connect = sqlite3.connect(f'{db_filename}')
                    cur = connect.cursor()
                    cur.execute("SELECT Предмети FROM Предмети")
                    res = cur.fetchall()
                    res = [item.lower() for tpl in res for item in tpl]
                    connect.close()
                    conn.close()
                    if infopred[0].lower() in res:
                        bot.send_message(message.chat.id, f"Такий предмет уже існує надішліть новий предмет")
                        bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)
                    else:
                        if infopred[1].isdigit():
                            if infopred[2].lower() not in ["екзамен","залік"]:

                                bot.send_message(message.chat.id,f"Тип підсумкового контролю вказаний неправильно")
                                bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)
                            else:
                                jurnal1_3_dodavanya_predmety_3(message, db_filename, infopred)
                        else:
                            bot.send_message(message.chat.id, f"Кількість кредитів ЕКТС вказано не правильно")
                            bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)

                else:
                    bot.send_message(message.chat.id, f"Такого предмету у нашій базі немає введіть новий предмет")
                    bot.register_next_step_handler(message, jurnal1_3_dodavanya_predmety_2, db_filename)
def jurnal1_3_dodavanya_predmety_3(message, db_filename,infopred):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    subject_name, credits, exam_type = infopred[0], infopred[1], infopred[2]
    cursor.execute("INSERT INTO Предмети (Предмети, Кредити_ЄКТС, Форма_підсумкового_контролю) VALUES (?, ?, ?)",(subject_name, credits, exam_type))
    conn.commit()
    conn.close()
    bot.send_message(message.chat.id, "Створюю таблиці з вашим предметом...")
    jurnal1_3_dodavanya_predmety_4_сreate(message, db_filename, subject_name, credits, exam_type)
def jurnal1_3_dodavanya_predmety_4_сreate(message, db_filename,subject_name, credits, exam_type):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    pre = subject_name
    pre = pre.replace(" ", "_")
    cursor.execute("SELECT Студенти FROM STUDENTY")
    students = cursor.fetchall()
    for i in range(1):
        cursor.execute(f'CREATE TABLE "{pre}_1" ( Студенти TEXT, модуль_1 TEXT, Н TEXT, [тема 0] TEXT,[тема 1] TEXT, [тема 2] TEXT, [тема 3] TEXT, [тема 4] TEXT, [тема 5] TEXT, [тема 6] TEXT, [тема 7] TEXT,  [тема 8] TEXT, [тема 9] TEXT)')
        cursor.execute(f'INSERT INTO "{pre}_1" ("Студенти", модуль_1, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Дата", NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')
        cursor.execute(f'INSERT INTO "{pre}_1" ("Студенти", модуль_1, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Тип заняття", NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')

        cursor.execute( f'CREATE TABLE "{pre}_2" ( Студенти TEXT, модуль_2 TEXT, Н TEXT, [тема 0] TEXT,[тема 1] TEXT, [тема 2] TEXT, [тема 3] TEXT, [тема 4] TEXT, [тема 5] TEXT, [тема 6] TEXT, [тема 7] TEXT,  [тема 8] TEXT, [тема 9] TEXT)')
        cursor.execute(f'INSERT INTO "{pre}_2" ("Студенти", модуль_2, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Дата",NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')
        cursor.execute(f'INSERT INTO "{pre}_2" ("Студенти", модуль_2, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Тип заняття", NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')

        cursor.execute(f'CREATE TABLE "{pre}_3" ("Студенти" TEXT,Індз TEXT,[{exam_type}] TEXT, модуль_1 TEXT, модуль_2 TEXT, Загальна_кількість_балів TEXT)')


        cursor.execute(f'CREATE TABLE "{pre}_Індивідуальні_години" (Студенти TEXT)')
        cursor.execute(f'INSERT INTO "{pre}_Індивідуальні_години" (Студенти) VALUES ("Дата")')
        cursor.execute(f'INSERT INTO "{pre}_Індивідуальні_години" (Студенти) VALUES ("Тип заняття")')

        cursor.execute(f'CREATE TABLE "{pre}_Студенти" ("Студенти" TEXT, модуль_1 TEXT, модуль_2 TEXT, [Індз] TEXT,[{exam_type}] TEXT, Загальна_кількість_балів TEXT)')
    # cursor.execute(f'INSERT INTO "{pre}_Студенти" ("Студенти", [модуль 1], [модуль 2], [індз], [підсумковий контроль], [загальна кількість балів]) VALUES ("Дата", NULL, NULL, NULL, NULL, NULL)')

    for student in students:
        cursor.execute(f"INSERT INTO \"{pre}_1\" (\"Студенти\") VALUES (?)", (str(student[0]),))
        cursor.execute(f"INSERT INTO \"{pre}_2\" (\"Студенти\") VALUES (?)", (str(student[0]),))
        cursor.execute(f"INSERT INTO \"{pre}_3\" (\"Студенти\") VALUES (?)", (str(student[0]),))
        cursor.execute(f"INSERT INTO \"{pre}_Студенти\" (\"Студенти\") VALUES (?)", (str(student[0]),))
        cursor.execute(f"INSERT INTO \"{pre}_Індивідуальні_години\" (\"Студенти\") VALUES (?)", (str(student[0]),))


    conn.commit()
    conn.close()
    bot.send_message(message.chat.id, "Таблиці успішно створенні✅")
    jurnal1_1(message)
def jurnal1_4(message,db_filename,user_grypa,subject):
    table = message.text

    subject = subject.replace(" ", "_")



    if table == '🔙Назад':
        conn = sqlite3.connect(f'{db_filename}')
        cursor = conn.cursor()
        cursor.execute('SELECT Предмети FROM Предмети')
        result = cursor.fetchall()

        keyboard = []
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        for row in result:
            subject = row[0]
            keyboard.append(subject)
            markup.add(types.KeyboardButton(subject))

        conn.close()
        bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
        bot.register_next_step_handler(message, jurnal1_3, db_filename, user_grypa, keyboard)

    elif table == 'ІНДЗ':
        indz(message,db_filename,subject)

    elif table.startswith('/'):
        bot.send_message(message.chat.id, "Ви ввели команду будь ласка оберіть розділ, який хочете редагувати")
        jurnal1_3(message, db_filename, user_grypa,subject)



    elif table in ['Модуль 1', 'Модуль 2']:
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Закритий_модуль_1, Закритий_модуль_2 FROM Предмети WHERE Предмети = '{subject.replace('_', ' ')}'")
        row = cursor.fetchone()
        if table == 'Модуль 1' and row[0]== 'Закритий модуль 1':
            bot.send_message(message.chat.id, f'Модуль 1 із предмету {subject.replace("_", " ")} закритий вашим викладачем ви вже не можете його редагувати, але можете переглядати оцінки своїх одогрупників')
            if table == 'Модуль 1':
                table = '1'
            elif table == 'Модуль 2':
                table = '2'

            jurnal_prerglad_ocinok_dla_starost(message,db_filename,user_grypa,subject, table)


        elif table == 'Модуль 2' and row[1]== 'Закритий модуль 2':
            bot.send_message(message.chat.id,f'Модуль 2 із предмету {subject.replace("_", " ")} закритий вашим викладачем ви вже не можете його редагувати, але можете переглядати оцінки своїх одогрупників')
            if table == 'Модуль 1':
                table = '1'
            elif table == 'Модуль 2':
                table = '2'

            jurnal_prerglad_ocinok_dla_starost(message, db_filename, user_grypa, subject, table)

        else:
            if table == 'Модуль 1':
                table = '1'
            elif table == 'Модуль 2':
                table = '2'

            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item1 = types.KeyboardButton('Додати оцінку')#
            item7 = types.KeyboardButton('Перегляд оцінок')#
            item2 = types.KeyboardButton('Редагувати назву теми')#
            item3 = types.KeyboardButton('Додати тему')#
            back = types.KeyboardButton('🔙Назад')
            markup.add(back)
            markup.add(item7)
            markup.add(item1)
            markup.add(item2)
            markup.add(item3)

            bot.send_message(message.chat.id, "Оберіть, що саме ви хочете редагувати в журналі:", reply_markup=markup)


            bot.register_next_step_handler(message, jurnal1_5, db_filename, user_grypa, subject, table)

    else:
        bot.send_message(message.chat.id, 'Такого варіанту немає')
        bot.register_next_step_handler(message,jurnal1_4, db_filename, user_grypa, subject)
def indz(message,db_filename,subject):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    #reply_markup = telebot.types.ReplyKeyboardRemove()
    cursor.execute("SELECT Студенти FROM STUDENTY")
    students = cursor.fetchall()

    # Формуємо повідомлення зі списком студентів
    students_list = "\n".join([student[0] + ' - 0' for student in students])
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('🔙Назад'))

    bot.send_message(message.chat.id,"Будь ласка надішліть мені список студентів вашої групи їхнє повне ім\'я.\nдля виставлення оцінок за ІНДЗ ️\nТакож я надішлю вам список вашої групи, для зручнішого виставлення оцінок")
    bot.send_message(message.chat.id,"ПІБ  Одногрупника - оцінка\nПІБ  Одногрупника - оцінка\nПІБ Одногрупника  - оцінка",reply_markup=markup)
    bot.send_message(message.chat.id, f"<code>{students_list}</code>", parse_mode=ParseMode.HTML)
    bot.register_next_step_handler(message, indz_2, db_filename, subject)
def indz_2(message,db_filename,subject):

    text = message.text
    if text == '🔙Назад':
        i = db_filename.split("-")
        user_grypa = i[0]
        jurnal1_2_1interval(message, db_filename, user_grypa)
    else:
        subject = subject.replace(" ", "_")
        table_name = subject + '_1'
        table_name_2 = subject + '_3'
        table_name_STUD = subject + '_Студенти'
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        rows = text.split("\n")
        split = []
        splitnot = []
        for row in rows:
            student_data = row.split(" - ")

            if len(student_data) == 2:
                split.append(row)

            else:
                splitnot.append(row)

        if len(splitnot) == 0:
            for row in split:
                student_data = row.split(" - ")
                name, grade = student_data
                cursor.execute(
                    f"UPDATE {table_name_2} SET Індз = ? WHERE Студенти = ?",
                    (grade, name))
                cursor.execute(f"UPDATE {table_name_STUD} SET Індз = ? WHERE Студенти = ?",
                    (grade, name))

            conn.commit()
            conn.close()
            bot.send_message(message.chat.id, "Дані успішно додано до бази даних.")
            Jurnal_1_5.jurnal1_5_dodavanna_ocinok(message,db_filename, table_name, subject)

        elif len(splitnot) > 0:
            list = tuple(splitnot)
            message_text = "\n\n".join(list)
            bot.send_message(message.chat.id, f"{message_text} Ось ці рядки я не зміг розпізнати, будь ласка надішліть ще раз і правильно за таким зразком\n\nПІБ - Оцінка\nПІБ - Оцінка")
            bot.register_next_step_handler(message, indz_2, db_filename,subject)
def jurnal1_5(message, db_filename, user_grypa, subject, table):
    if message.text == '🔙Назад':
        conn = sqlite3.connect(f'{db_filename}')
        cursor = conn.cursor()
        cursor.execute('SELECT Предмети FROM Предмети')
        result = cursor.fetchall()

        keyboard = []
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        for row in result:
            subject = row[0]
            keyboard.append(subject)
            markup.add(types.KeyboardButton(subject))

        conn.close()
        bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
        bot.register_next_step_handler(message, jurnal1_3, db_filename, user_grypa, keyboard)

    elif message.text == 'Додати оцінку':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Закритий_предмет FROM Предмети WHERE Предмети = '{subject.replace('_', ' ')}'")
        row = cursor.fetchone()
        row = row[0]
        if row == 'Закритий предмет':
            bot.send_message(message.chat.id,f'Журнал предмету {subject.replace("_", " ")} закритий вашим викладачем ви вже не можете його редагувати)')
            jurnal1_1(message)

        else:
            # Формуємо назву таблиці у форматі {subject}_{table}
            table_name = f'{subject}_{table}'

            # Отримуємо назви стовпців таблиці
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns if column[1] != 'Студенти' and column[1] != 'модуль_1' and column[1] != 'модуль_2' and column[1] != 'Н']

            # Створюємо клавіатуру з кнопками зі списку column_names
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for column_name in column_names:
                markup.add(column_name)


            # Надсилаємо повідомлення з клавіатурою
            bot.send_message(message.chat.id, "Оберіть тему в яку хочете внести оцінки:", reply_markup=markup)
            bot.register_next_step_handler(message, Jurnal_1_5.jurnal1_5_1, db_filename, user_grypa, table_name ,column_names,subject,table)
            conn.close()#1_5_1

    elif message.text == 'Редагувати назву теми':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Закритий_предмет FROM Предмети WHERE Предмети = '{subject.replace('_', ' ')}'")
        row = cursor.fetchone()
        row = row[0]
        if row == 'Закритий предмет':
            bot.send_message(message.chat.id,
                             f'Журнал предмету {subject.replace("_", " ")} закритий вашим викладачем ви вже не можете його редагувати)')
            jurnal1_1(message)

        else:

            # Формуємо назву таблиці у форматі {subject}_{table}
            table_name = f'{subject}_{table}'

            # Отримуємо назви стовпців таблиці
            cursor.execute(f"PRAGMA table_info({table_name})")
            columns = cursor.fetchall()
            column_names = [column[1] for column in columns if column[1] != 'Студенти' and column[1] != 'модуль_1' and column[1] != 'модуль_2' and column[1] != 'Н']


            # Створюємо клавіатуру з кнопками зі списку column_names
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for column_name in column_names:
                markup.add(column_name)


            # Надсилаємо повідомлення з клавіатурою
            bot.send_message(message.chat.id, "Оберіть тему, яку ви хочете відредагувати:", reply_markup=markup)
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_1, db_filename, user_grypa, subject, table, column_names)

            conn.close()

    elif message.text == 'Додати тему':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Закритий_предмет FROM Предмети WHERE Предмети = '{subject.replace('_', ' ')}'")
        row = cursor.fetchone()
        row = row[0]
        if row == 'Закритий предмет':
            bot.send_message(message.chat.id,f'Журнал предмету {subject.replace("_", " ")} закритий вашим викладачем ви вже не можете його редагувати)')
            jurnal1_1(message)

        else:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            bot.send_message(message.chat.id,"Надішліть назву нової теми для предмету", reply_markup=markup)
            bot.register_next_step_handler(message,jurnal_1_dodavanna_temy, db_filename,subject, table)

    elif message.text == 'Перегляд оцінок':
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('1')
        item2 = types.KeyboardButton('2')
        markup.add(item1)
        markup.add(item2)
        bot.send_message(message.chat.id, "1 - переглянути оцінки для окремого студента з всього предмету\n2 - Перегляд оцінок всієї групи із окремої теми".format(message.from_user), reply_markup=markup)
        bot.register_next_step_handler(message,  look_grade_1, db_filename, user_grypa, subject, table)

    else:
        bot.send_message(message.chat.id,"Ви ввели щось не то")

        bot.register_next_step_handler(message, jurnal1_5, db_filename, user_grypa, subject, table)
def look_grade_1(message,db_filename, user_grypa, subject, table):
    text = message.text
    if text == "2":
        look_grade_2(message, db_filename, user_grypa, subject, table)
    elif text == '1':
        bot.send_message(message.chat.id,"Надішліть повне ПІБ студента у якого хочете переглянути оцінки")
        bot.register_next_step_handler(message,look_grade_student_1,db_filename, user_grypa, subject, table)
def look_grade_student_1(message,db_filename, user_grypa, subject, table):
    text = message.text
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute("SELECT Студенти FROM STUDENTY")
    students = cursor.fetchall()
    students = [item for tpl in students for item in tpl]
    if text not in students:
        bot.send_message(message.chat.id, "Такого Студента немає у вашій групі будь ласка введіть ще раз повне ПІБ")
        bot.register_next_step_handler(message, look_grade_student_1, db_filename, user_grypa, subject, table)
    else:
        look_grade_student_2(message,db_filename, user_grypa, subject, table, text)
def look_grade_student_2(message,db_filename, user_grypa, subject, table, text):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    table_name = f'{subject}_{table}'
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = cursor.fetchall()
    column_names = [column[1] for column in columns if
                    column[1] != 'Студенти']
    column_names_str = ', '.join([f'`{column}`' for column in column_names])
    cursor.execute(f"SELECT {column_names_str} FROM {table_name} WHERE Студенти = '{text}'")
    columns = cursor.fetchall()
    updated_list = [[0 if item is None else item for item in tuple_item] for tuple_item in columns]
    # Об'єднаємо всі значення у один список
    flat_list = [item for tuple_item in updated_list for item in tuple_item]
    messagee = ''
    for k, i in zip(flat_list, column_names):
        if i == "Н":
            messagee += f"Кількість Н - {k}\n"
        else:
            messagee += f"{i} - {k}\n"
    bot.send_message(message.chat.id, f"Оцінки студента {text}, з предмету {subject} за модуль {table}\n{messagee}")
    jurnal1_2(message, db_filename, user_grypa)
def look_grade_2(message,db_filename, user_grypa, subject, table):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()

    # Формуємо назву таблиці у форматі {subject}_{table}
    table_name = f'{subject}_{table}'

    # Отримуємо назви стовпців таблиці
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = cursor.fetchall()
    column_names = [column[1] for column in columns if
                    column[1] != 'Студенти' and column[1] != 'Н']

    # Створюємо клавіатуру з кнопками зі списку column_names
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('🔙Назад'))
    for column_name in column_names:
        markup.add(column_name)

    bot.send_message(message.chat.id,f"Оберіть тему, з модуля {table} з якої ви хочете переглянути оцінки своїх одногрупників або зразу з цілого модуля:",reply_markup=markup)
    bot.register_next_step_handler(message, jurnal_1_pereglad_ocinok, db_filename, user_grypa, subject, table,column_names)
def jurnal_1_pereglad_ocinok(message, db_filename, user_grypa, subject, table,column_names):
    text = message.text

    if text == '🔙Назад':
        jurnal1_2(message, db_filename, user_grypa)
    elif text not in column_names:
        bot.send_message(message.chat.id,"Ви вибрали не вірну тему оберіть ще раз")
        bot.register_next_step_handler(message, jurnal_1_pereglad_ocinok, db_filename, user_grypa, subject, table,column_names)
    else:
        table_name = f'{subject}_{table}'
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()
        # Формуємо повідомлення зі списком студентів
        students_list = "\n".join([student[0] + ' - ' for student in students])

        cursor.execute(f"SELECT [{text}] FROM {table_name} ")
        ocinky = cursor.fetchall()
        ocinky = ocinky[2:]
        ocinky = [item for tpl in ocinky for item in tpl]
        ocinky2 = []
        for i in ocinky:
            if i == None:
                ocinky2.append("0")
            else:
                ocinky2.append(i)

        results = []
        students_list1 = students_list.split("\n")
        for k, i in enumerate(students_list1):
            results.append(i + ocinky2[k])


        gem = "\n".join([row for row in results])

        bot.send_message(message.chat.id, f"Ось оцінки із предмету {subject.replace('_', ' ')} із теми {text} із модуля {table}:\n\n{gem}")
        bot.register_next_step_handler(message, jurnal_1_pereglad_ocinok, db_filename, user_grypa, subject, table,column_names)
def jurnal_1_dodavanna_temy(message, db_filename, subject, table):
    new_tema = message.text
    kay = ["модуль 1", 'модуль 2', 'Модуль 1', 'Модуль 2', "залік", "Залік", 'Екзамен', 'екзамен',
           'Підсумковий контроль', 'підсумковий контроль','студенти','Студенти',]
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()
    cursor.execute(f"PRAGMA table_info({subject}_{table})")
    columns = cursor.fetchall()
    columns_names = [item[1] for item in columns]
    if new_tema == '🔙Назад':
        if message.text == '🔙Назад':
            conn = sqlite3.connect(f'{db_filename}')
            cursor = conn.cursor()
            cursor.execute('SELECT Предмети FROM Предмети')
            result = cursor.fetchall()

            keyboard = []
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for row in result:
                subject = row[0]
                keyboard.append(subject)
                markup.add(types.KeyboardButton(subject))

            conn.close()
            user_grypa = db_filename.split("-")
            user_grypa = user_grypa[0]
            bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
            bot.register_next_step_handler(message, jurnal1_3, db_filename, user_grypa, keyboard)
    elif new_tema in columns_names:
        bot.send_message(message.chat.id, f'Така тема вже існує ви не можете додати її: {new_tema}\nНадішліть мені іншу назву теми')
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy, db_filename, subject, table)
    elif new_tema in kay:
        bot.send_message(message.chat.id, f'Таку тему додати неможливо: {new_tema}\nНадішліть мені іншу назву теми')
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy, db_filename, subject, table)


    else:
        # Підключення до бази даних
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()

        # Створення нової колонки
        query = f"ALTER TABLE {subject}_{table} ADD COLUMN [{new_tema}] TEXT"
        cursor.execute(query)

        # Збереження змін у базі даних
        conn.commit()
        conn.close()
        jurnal_1_dodavanna_temy_2(message, db_filename, subject, table, new_tema)
def jurnal_1_dodavanna_temy_2(message, db_filename, subject, table, new_tema):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    info = types.KeyboardButton('Семінар')
    info2 = types.KeyboardButton('Практична')
    info3 = types.KeyboardButton('Лабораторна')
    markup.add(info)
    markup.add(info2)
    markup.add(info3)
    bot.send_message(message.chat.id, "Оберіть тип заняття із кнопок нище", reply_markup=markup)
    bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_3, db_filename, subject, table,new_tema)
def jurnal_1_dodavanna_temy_3(message, db_filename, subject, table,new_tema):
    text = message.text
    if text == 'Практична' or text == 'Семінар' or text == 'Лабораторна':
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        text2 = "Тип заняття"
        cursor.execute(f"UPDATE {subject}_{table} SET [{new_tema}] = ? WHERE Студенти = ?", (text,text2))
        conn.commit()
        conn.close()

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        info = types.KeyboardButton('Актуальна дата')
        info2 = types.KeyboardButton('Ввести дату')
        markup.add(info)
        markup.add(info2)
        bot.send_message(message.chat.id, "Оберіть варіант дати",reply_markup=markup)
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_4, db_filename, subject, table, new_tema)

    else:
        bot.send_message(message.chat.id, "Такого варіанту немає оберіть ще раз тип заняття")
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_3, db_filename, subject, table, new_tema)
def jurnal_1_dodavanna_temy_4(message, db_filename, subject, table,new_tema):
    text = message.text
    user_id = message.chat.id
    if text == 'Актуальна дата':
        current_date = datetime.datetime.now()
        formatted_date = current_date.strftime("%d.%m.%Y")
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        text2 = "Дата"
        cursor.execute(f"UPDATE {subject}_{table} SET [{new_tema}] = ? WHERE Студенти = ?", (formatted_date, text2))
        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, f"Дата додана {formatted_date} до теми {new_tema}")
        connect = sqlite3.connect('users.db')
        cursor1 = connect.cursor()
        cursor1.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
        user_rol = cursor1.fetchone()
        user_rol = user_rol[0]
        conn.close()
        connect.close()
        if user_rol == 'староста':
            jurnal1_1(message)
        elif user_rol == 'викладач':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item2 = types.KeyboardButton('Робота з журналами')
            homework = types.KeyboardButton('Додати домашнє')
            item4 = types.KeyboardButton('Оголошення для групи')
            back = types.KeyboardButton('🔙Назад')
            markup.add(back)
            markup.add(item2)
            markup.add(homework)
            markup.add(item4)
            bot.send_message(message.chat.id,
                             "Ви попали в меню викладачам оберіть функцію з якою хочете працювати".format(
                                 message.from_user), reply_markup=markup)
            bot.register_next_step_handler(message, menu_vikladacham_2)

    elif text == 'Ввести дату':
        bot.send_message(message.chat.id, "Введіть дату за зразком 01.01.2023")
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_5, db_filename, subject, table, new_tema)
    else:
        bot.send_message(message.chat.id, "Такого варінту немає оберіть тип дати ще раз")
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_4, db_filename, subject, table, new_tema)
def jurnal_1_dodavanna_temy_5(message, db_filename, subject, table,new_tema):
    text = message.text
    lines = text.split(".")
    notcifra = []
    user_id = message.chat.id
    i = 0
    if len(lines) > 3:
        bot.send_message(message.chat.id, "Не правлений формат рядка\nВведіть дату за зразком 01.01.2023")
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_5, db_filename, subject, table, new_tema)
    elif text.startswith('/'):
        bot.send_message(message.chat.id, f"Ви надіслали команду, а не дату, яку маєте обов'язково додати")
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_5, db_filename, subject, table, new_tema)
    else:

        for item in lines:
            if item.isdigit():
                i+=1
            else:
                notcifra.append(item)

    if len(notcifra)>0:
        bot.send_message(message.chat.id, "Не правлений формат рядка\nВведіть дату за зразком 01.01.2023")
        bot.register_next_step_handler(message, jurnal_1_dodavanna_temy_5, db_filename, subject, table, new_tema)
    else:
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        text2 = "Дата"
        cursor.execute(f"UPDATE {subject}_{table} SET [{new_tema}] = ? WHERE Студенти = ?", (text, text2))
        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, f"Дата додана {text} до теми {new_tema}")

        connect = sqlite3.connect('users.db')
        cursor1 = connect.cursor()
        cursor1.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
        user_rol = cursor1.fetchone()
        user_rol = user_rol[0]
        conn.close()
        connect.close()
        if user_rol == 'староста':
            jurnal1_1(message)
        elif user_rol == 'викладач':
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            item2 = types.KeyboardButton('Робота з журналами')
            homework = types.KeyboardButton('Додати домашнє')
            item4 = types.KeyboardButton('Оголошення для групи')
            back = types.KeyboardButton('🔙Назад')
            markup.add(back)
            markup.add(item2)
            markup.add(homework)
            markup.add(item4)
            bot.send_message(message.chat.id,
                             "Ви попали в меню викладачам оберіть функцію з якою хочете працювати".format(
                                 message.from_user), reply_markup=markup)
            bot.register_next_step_handler(message, menu_vikladacham_2)


class EditTEMA:
    def __init__(self, bot):
        self.bot = bot
        self.tema = None

    @staticmethod
    def jurnal1_tema_1(message, db_filename, user_grypa, subject, table, column_names):
        tema = message.text

        if tema == '🔙Назад':
            conn = sqlite3.connect(f'{db_filename}')
            cursor = conn.cursor()
            cursor.execute('SELECT Предмети FROM Предмети')
            result = cursor.fetchall()

            keyboard = []
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for row in result:
                subject = row[0]
                keyboard.append(subject)
                markup.add(types.KeyboardButton(subject))

            conn.close()
            bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
            bot.register_next_step_handler(message, EditTEMA.jurnal1_3, db_filename, user_grypa, keyboard)
        elif tema.startswith('/'):
            bot.send_message(message.chat.id,
                             f"ви надіслали команду, а не назву теми, яку хочете редагувати оберіть будь ласка ще раз тему")
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_1, db_filename, user_grypa, subject, table,
                                           column_names)
        else:
            if tema in column_names:
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                markup.add(types.KeyboardButton('🔙Назад'))
                bot.send_message(message.chat.id, f"Введіть нову назву теми, яку ви вибрали {tema}",
                                 reply_markup=markup)
                bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_2, db_filename, user_grypa, subject,
                                               table, column_names, tema)
            else:
                bot.send_message(message.chat.id,
                                 f"Ви вибрали не вірну тему {tema} оберіть будь ласка ще раз і правильну тему ")
                bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_1, db_filename, user_grypa, subject,
                                               table, column_names)

    @staticmethod
    def jurnal1_tema_2(message, db_filename, user_grypa, subject, table,column_names,tema):
        tema_new = message.text
        kay = ["модуль 1", 'модуль 2', 'Модуль 1', 'Модуль 2', "залік", "Залік", 'Екзамен', 'екзамен',
               'Підсумковий контроль', 'підсумковий контроль','студенти','Студенти']
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({subject}_{table})")
        columns = cursor.fetchall()
        names = [item[1] for item in columns]


        if tema_new == '🔙Назад':
            conn = sqlite3.connect(f'{db_filename}')
            cursor = conn.cursor()
            cursor.execute('SELECT Предмети FROM Предмети')
            result = cursor.fetchall()

            keyboard = []
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for row in result:
                subject = row[0]
                keyboard.append(subject)
                markup.add(types.KeyboardButton(subject))

            conn.close()
            bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
            bot.register_next_step_handler(message, jurnal1_3, db_filename, user_grypa, keyboard)
        elif tema_new in names:
            bot.send_message(message.chat.id, f'Така тема вже існує ви не можете додати 2 одинакових теми‼️: {tema_new}\nНадішліть мені іншу назву теми')
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_2, db_filename,user_grypa, subject, table,column_names,tema)
        elif tema_new in kay:
            bot.send_message(message.chat.id, f'Таку назву теми неможливо встановити: {tema_new}\nНадішліть мені іншу назву теми')
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_2, db_filename, user_grypa, subject, table,column_names,tema)
        elif tema_new.startswith('/'):
            bot.send_message(message.chat.id, f"ви надіслали команду а не назву теми будь ласка введіть ще раз")
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_2 , db_filename, user_grypa, subject, table,column_names,tema)
        else:


            table_name = f"{subject}_{table}"

            conn = sqlite3.connect(db_filename)
            cursor = conn.cursor()
            cursor.execute(f"ALTER TABLE {table_name} RENAME COLUMN '{tema}' TO '{tema_new}'")
            conn.commit()
            conn.close()
            bot.send_message(message.chat.id, f"Тему під назвою {tema} успішно переіменовано на {tema_new}\nТепер додамо тип заняття")
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            info = types.KeyboardButton('Семінар')
            info2 = types.KeyboardButton('Практична')
            info3 = types.KeyboardButton('Лабораторна')
            markup.add(info)
            markup.add(info2)
            markup.add(info3)
            bot.send_message(message.chat.id, "Оберіть тип заняття із кнопок нище", reply_markup=markup)
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_3, db_filename, user_grypa, subject, table, column_names, tema_new)
    @staticmethod
    def jurnal1_tema_3(message, db_filename, user_grypa, subject, table,column_names,tema_new):
        text = message.text
        if text == 'Практична' or text == 'Семінар' or text == 'Лабораторна':
            conn = sqlite3.connect(db_filename)
            cursor = conn.cursor()
            text2 = "Тип заняття"
            cursor.execute(f"UPDATE {subject}_{table} SET [{tema_new}] = ? WHERE Студенти = ?", (text,text2))
            conn.commit()
            conn.close()
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            info = types.KeyboardButton('Актуальна дата')
            info2 = types.KeyboardButton('Ввести дату')
            markup.add(info)
            markup.add(info2)
            bot.send_message(message.chat.id, "Оберіть варіант дати", reply_markup=markup)
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_4, db_filename, user_grypa, subject, table, column_names,tema_new)
        else:
            bot.send_message(message.chat.id, "Такого варіанту немає будь ласка виберіть ти заняття ще раз")
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_3, db_filename, user_grypa, subject, table,column_names,tema_new)
    @staticmethod
    def jurnal1_tema_4(message, db_filename, user_grypa, subject, table,column_names,tema_new):
        text = message.text
        user_id = message.chat.id
        if text == 'Актуальна дата':
            current_date = datetime.datetime.now()
            formatted_date = current_date.strftime("%d.%m.%Y")
            conn = sqlite3.connect(db_filename)
            cursor = conn.cursor()
            text2 = "Дата"
            cursor.execute(f"UPDATE {subject}_{table} SET [{tema_new}] = ? WHERE Студенти = ?", (formatted_date, text2))
            conn.commit()
            conn.close()
            bot.send_message(message.chat.id, f"Дата додана {formatted_date} до теми {tema_new}")
            connect = sqlite3.connect('users.db')
            cursor1 = connect.cursor()
            cursor1.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
            user_rol = cursor1.fetchone()
            user_rol = user_rol[0]
            conn.close()
            connect.close()
            if user_rol == 'староста':
                jurnal1_1(message)
            elif user_rol == 'викладач':
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item2 = types.KeyboardButton('Робота з журналами')
                homework = types.KeyboardButton('Додати домашнє')
                item4 = types.KeyboardButton('Оголошення для групи')
                back = types.KeyboardButton('🔙Назад')
                markup.add(back)
                markup.add(item2)
                markup.add(homework)
                markup.add(item4)
                bot.send_message(message.chat.id,
                                 "Ви попали в меню викладачам оберіть функцію з якою хочете працювати".format(
                                     message.from_user), reply_markup=markup)
                bot.register_next_step_handler(message, menu_vikladacham_2)

        elif text == 'Ввести дату':
            bot.send_message(message.chat.id, "Введіть дату за зразком 01.01.2023")
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_5, db_filename, user_grypa, subject, table,column_names,tema_new)
        else:
            bot.send_message(message.chat.id, "Такого варінту немає оберіть тип дати ще раз")
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_4, db_filename, user_grypa, subject, table,column_names,tema_new)
    @staticmethod
    def jurnal1_tema_5(message, db_filename, user_grypa, subject, table,column_names,tema_new):
        text = message.text
        lines = text.split(".")
        notcifra = []
        user_id = message.chat.id
        i = 0
        if len(lines) > 3:
            bot.send_message(message.chat.id, "Не правлений формат рядка\nВведіть дату за зразком 01.01.2023")
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_5, db_filename, user_grypa, subject, table,column_names,tema_new)
        elif text.startswith('/'):
            bot.send_message(message.chat.id, f"Ви надіслали команду, а не дату, яку маєте обов'язково додати")
            bot.register_next_step_handler(message, EditTEMA.jurnal1_tema_5, db_filename, user_grypa, subject, table,column_names,tema_new)
        else:

            for item in lines:
                if item.isdigit():
                    i += 1
                else:
                    notcifra.append(item)

        if len(notcifra) > 0:
            bot.send_message(message.chat.id, "Не правлений формат рядка\nВведіть дату за зразком 01.01.2023")
            bot.register_next_step_handler(message,  EditTEMA.jurnal1_tema_5, db_filename, user_grypa, subject, table,column_names,tema_new)
        else:
            conn = sqlite3.connect(db_filename)
            cursor = conn.cursor()
            text2 = "Дата"
            cursor.execute(f"UPDATE {subject}_{table} SET [{tema_new}] = ? WHERE Студенти = ?", (text, text2))
            bot.send_message(message.chat.id, f"Дата додана {text} до теми {tema_new}")
            connect = sqlite3.connect('users.db')
            cursor1 = connect.cursor()
            cursor1.execute(f"SELECT roli FROM login_id WHERE id = {user_id}")
            user_rol = cursor1.fetchone()
            user_rol = user_rol[0]
            conn.close()
            connect.close()
            if user_rol == 'староста':
                jurnal1_1(message)
            elif user_rol == 'викладач':
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                item2 = types.KeyboardButton('Робота з журналами')
                homework = types.KeyboardButton('Додати домашнє')
                item4 = types.KeyboardButton('Оголошення для групи')
                back = types.KeyboardButton('🔙Назад')
                markup.add(back)
                markup.add(item2)
                markup.add(homework)
                markup.add(item4)
                bot.send_message(message.chat.id,
                                 "Ви попали в меню викладачам оберіть функцію з якою хочете працювати".format(
                                     message.from_user), reply_markup=markup)
                bot.register_next_step_handler(message, menu_vikladacham_2)

class Jurnal_1_5:
    def __init__(self, bot):
        self.bot = bot

    @staticmethod
    def jurnal1_5_1(message, db_filename, user_grypa,  table_name,column_names,subject,table):
        tema = message.text
        if tema == '🔙Назад':
            conn = sqlite3.connect(f'{db_filename}')
            cursor = conn.cursor()
            cursor.execute('SELECT Предмети FROM Предмети')
            result = cursor.fetchall()

            keyboard = []
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for row in result:
                subject = row[0]
                keyboard.append(subject)
                markup.add(types.KeyboardButton(subject))

            conn.close()
            bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
            bot.register_next_step_handler(message, jurnal1_3, db_filename, user_grypa, keyboard)

        elif tema in column_names:
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            bot.send_message(message.chat.id,"Будь ласка надішліть мені список студентів вашої групи їхнє повне ім\'я.\nЗа таким зразком⬇️\nТакож я надішлю вам список вашої групи, для зручнішого виставлення оціно\n\n‼️Зауважте я надіслав вам зразу актуальні оцінки студента ви можете їх змінити при потребі, але не рекомендуємо це робити якщо до того оцінки були виставлені‼️\n\nПІБ(Одногрупника) - оцінка або 'Н'\nПІБ(Одногрупника) - оцінка або 'Н'\nПІБ(Одногрупника) - оцінка або 'Н'", reply_markup=markup)

            conn = sqlite3.connect(db_filename)
            cursor = conn.cursor()

            # Витягуємо всі значення зі стовпця "Студенти" таблиці "STUDENTY"
            cursor.execute("SELECT Студенти FROM STUDENTY")
            students = cursor.fetchall()
            students = [row[0] for row in students]
            grates = []
            for stud in students:
                cursor.execute(f"SELECT [{tema}] FROM {table_name} WHERE Студенти = ?", (stud,))
                grate = cursor.fetchone()
                grate = grate[0]
                if grate == None:
                    grate = 0
                    grates.append(grate)
                else:
                    grates.append(grate)

            # Формуємо повідомлення зі списком студентів
            students_list = "\n".join([f"{student} - {grate}" for student, grate in zip(students, grates)])
            # Надсилаємо повідомлення зі списком студентів у бота
            bot.send_message(message.chat.id, f"<code>{students_list}</code>", parse_mode=ParseMode.HTML)
            conn.close()
            bot.register_next_step_handler(message,Jurnal_1_5.jurnal1_5_2, db_filename, user_grypa,  table_name,tema,subject,table)
    @staticmethod
    def jurnal1_5_2(message, db_filename, user_grypa, table_name, tema,subject,table):
        text = message.text

        if text == '🔙Назад':
            conn = sqlite3.connect(f'{db_filename}')
            cursor = conn.cursor()
            cursor.execute('SELECT Предмети FROM Предмети')
            result = cursor.fetchall()

            keyboard = []
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
            markup.add(types.KeyboardButton('🔙Назад'))
            for row in result:
                subject = row[0]
                keyboard.append(subject)
                markup.add(types.KeyboardButton(subject))

            conn.close()
            bot.send_message(message.chat.id, "Оберіть предмет з яким Ви хочете працювати", reply_markup=markup)
            bot.register_next_step_handler(message, jurnal1_3, db_filename, user_grypa, keyboard)
        else:
            conn = sqlite3.connect(db_filename)
            cursor = conn.cursor()
            rows = text.split("\n")
            split = []
            splitnot = []
            for row in rows:
                student_data = row.split(" - ")
                if len(student_data) == 2:
                    split.append(row)

                else:
                    splitnot.append(row)


            if len(splitnot) == 0:
                for row in split:
                    student_data = row.split(" - ")
                    name, grade = student_data
                    cursor.execute(
                        f"UPDATE {table_name} SET [{tema}] = ? WHERE Студенти = ?",
                        (grade, name))

                conn.commit()
                conn.close()
                bot.send_message(message.chat.id, "Дані успішно додано до бази даних.")
                Jurnal_1_5.jurnal1_5_dodavanna_ocinok(message, db_filename, table_name, subject)

            elif len(splitnot) > 0:
                list = tuple(splitnot)
                message_text = "\n\n".join(list)
                bot.send_message(message.chat.id, f"{message_text} Ось ці рядки я не зміг розпізнати, будь ласка надішліть ще раз і правильно за таким зразком\n\nПІБ - Оцінка\nПІБ - Оцінка")
                bot.register_next_step_handler(message, Jurnal_1_5.jurnal1_5_2, db_filename, user_grypa, table_name, tema, subject, table)

    @staticmethod
    def jurnal1_5_dodavanna_ocinok(message,db_filename, table_name, subject):
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({table_name})")
        columns = cursor.fetchall()
        column_names = column_names = [column[1] for column in columns if
                                       column[1] != 'Студенти' and column[1] != 'модуль_1' and column[1] != 'модуль_2' and
                                       column[1] != 'Н']

        modylchislo = table_name[-1]
        module = f"модуль_{modylchislo}"
        modyl = 0
        H = 0
        # Витягуємо всі значення зі стовпців тем для кожного студента
        students_topics = {}
        for column_name in column_names:
            cursor.execute(f"SELECT Студенти, [{column_name}] FROM {table_name}")
            results = cursor.fetchall()
            for row in results:
                student = row[0]
                topic_value = row[1]
                if student not in students_topics:
                    students_topics[student] = []
                students_topics[student].append(topic_value)

        students_topics.pop(next(iter(students_topics)))

        for i in students_topics:
            for k in range(len(students_topics[i])):
                if students_topics[i][k] == None:
                    continue
                elif students_topics[i][k] == 'н' or students_topics[i][k] == 'Н':
                    H += 1

                elif students_topics[i][k] in ['Лекція', 'Практична', 'Семінар', 'Лабораторні роботи']:
                    continue
                elif students_topics[i][k].isdigit():
                    students_topics[i][k] = int(students_topics[i][k])
                    modyl += students_topics[i][k]
            modyl = str(modyl)


            cursor.execute(f"UPDATE {table_name} SET {module} = ?, Н = ? WHERE Студенти = ?", (modyl, H, i))

            modyl = 0
            H = 0
        # Застосування змін до бази даних
        conn.commit()

        # Закриття курсора та з'єднання з базою даних
        cursor.close()
        conn.close()
        Jurnal_1_5.jurnal1_5_dodavanna_ocinok_v_inshy_table(message,db_filename, table_name, subject)
    @staticmethod
    def jurnal1_5_dodavanna_ocinok_v_inshy_table(message,db_filename,table_name, subject):
        subject = subject.replace(' ','_')
        pred = subject

        modylchislo = table_name[-1]
        module = f"модуль_{modylchislo}"


        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Студенти, {module} FROM {table_name}")
        OCIN =cursor.fetchall()
        OCIN.remove(OCIN[0])



        for i in OCIN:

            if len(i) == 2:
                name, grade = i
                cursor.execute(f"UPDATE {pred}_3 SET {module} = ? WHERE Студенти = ?",(grade,name))
                cursor.execute(f"UPDATE {pred}_Студенти SET {module} = ? WHERE Студенти = ?", (grade,name))

        conn.commit()
        conn.close()
        Jurnal_1_5.jurnal1_5_dodavanna_ocinok_v_inshy_table2(message,db_filename, table_name,pred)

    @staticmethod
    def jurnal1_5_dodavanna_ocinok_v_inshy_table2(message,db_filename,table_name,pred):
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"PRAGMA table_info({pred}_3)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if column[1] != 'Курсова(якщо_є)' and column[1] != 'Загальна_кількість_балів']
        columns_str = ', '.join(column_names)
        cursor.execute(f"SELECT {columns_str} FROM {pred}_3")
        results = cursor.fetchall()


        supisok = []
        for i in results:
            suma_ocin = 0
            for index, k in enumerate(i):

                if index == 0 or index == 1:  # Пропускаємо перший елемент
                    continue
                if k is None or k == '':
                    continue
                else:
                    suma_ocin += int(k)
            supisok.append((i[0], str(suma_ocin)))  # Додаємо кортеж з ім'ям і сумою
        for i in supisok:
            if len(i)== 2:
                name, grade = i
                cursor.execute(f"UPDATE {pred}_3 SET Загальна_кількість_балів = ? WHERE Студенти = ?", (grade, name))

        conn.commit()
        conn.close()
        Jurnal_1_5.jurnal1_5_dodavanna_ocinok_v_inshy_table2_1(message,db_filename, table_name,pred)

    @staticmethod
    def jurnal1_5_dodavanna_ocinok_v_inshy_table2_1(message,db_filename,table_name,pred):
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Студенти, Загальна_кількість_балів FROM {pred}_3")
        results = cursor.fetchall()



        for i in results:
            if len(i) == 2:
                name, grade = i
                cursor.execute(f"UPDATE {pred}_Студенти SET Загальна_кількість_балів = ? WHERE Студенти = ?",(grade, name))

        conn.commit()
        conn.close()

        Jurnal_1_5.jurnal1_5_dodavanna_ocinok_v_inshy_table2_2(message, db_filename, table_name, pred)
    @staticmethod
    def jurnal1_5_dodavanna_ocinok_v_inshy_table2_2(message, db_filename, table_name, pred):
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Студенти FROM STUDENTY")
        results = cursor.fetchall()
        results = [item for tpl in results for item in tpl]

        cursor.execute(f"PRAGMA table_info({pred}_3)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if
                        column[1] != 'Курсова(якщо_є)' and column[1] != 'Загальна_кількість_балів' and column[
                            1] != 'Студенти']
        columns_str = ', '.join(column_names)

        for i in results:
            cursor.execute(f"SELECT {columns_str} FROM {pred}_3 WHERE Студенти = ?", (i,))
            grate = cursor.fetchall()
            grate = [item for tpl in grate for item in tpl]

            sum = 0
            for j in grate:

                if j is not None:
                    sum += int(j)

            cursor.execute(f"UPDATE {pred}_3 SET Загальна_кількість_балів = ? WHERE Студенти = ?", (sum, i))
            cursor.execute(f"UPDATE {pred}_Студенти SET Загальна_кількість_балів = ? WHERE Студенти = ?", (sum, i))

        conn.commit()
        conn.close()
        message_handler_start(message)

class Create:
    def __init__(self, bot):
        self.bot = bot

    @staticmethod
    def create_table(message,db_filename, user_grypa):
        pred1 = message.text
        pred = pred1.replace(" ","_")
        gryp = user_grypa
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute(f"SELECT Студенти FROM STUDENTY")
        results = cursor.fetchall()
        results = [item for tpl in results for item in tpl]

        cursor.execute(f"SELECT Форма_підсумкового_контролю FROM Предмети WHERE Предмети = '{pred1}'")
        exam = cursor.fetchone()[0]


        # Відкриваємо існуючий Excel-файл
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        k = 5
        number = 1
        for i in results:
            sheet[f'A{k}'] = number
            sheet[f'B{k}'] = i
            sheet.row_dimensions[k].height = 20
            number += 1
            k += 1

        sheet.merge_cells(f'B{k}:C{k}')
        sheet[f'B{k}'] = 'Максимальна кількість балів'
        sheet.row_dimensions[k].height = 25
        sheet.row_dimensions[k + 1].height = 39
        sheet[f'B{k + 1}'] = "Підпис викладача"
        sheet.merge_cells(f'B{k + 1}:C{k + 1}')

        # Змінюємо значення в комірках
        sheet['A1'] = 'Група'
        sheet['B1'] = "Навчальна Дисципліна"
        sheet['C1'] = 'Форма підсумкового контролю'
        sheet['E1'] = 'Викладач'
        sheet['A2'] = gryp
        sheet['B2'] = pred.replace('_', ' ')
        sheet['C2'] = exam
        sheet.merge_cells('A3:C3')
        sheet['A4'] = "№ з/п"
        sheet['B4'] = "ПІБ Студентів"
        sheet['C4'] = "Дата"
        sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['A3'] = 'Вид контролю, тема'
        sheet.merge_cells('A3:C3')
        # Вирівнюємо текст по центру в комірці
        sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        sheet['C1'].alignment = openpyxl.styles.Alignment(wrapText=True)
        sheet['E1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['B2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['C2'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['A4'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['B4'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        sheet['C4'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        sheet.column_dimensions['B'].width = 32
        sheet.column_dimensions['C'].width = 14
        sheet.column_dimensions['A'].width = 7

        sheet.row_dimensions[1].height = 55
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[3].height = 115
        sheet.row_dimensions[4].height = 75

        cursor.execute(f"PRAGMA table_info({pred}_1)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if
                        column[1] != 'Студенти' and column[1] != 'Н' and column[1] != 'модуль_1' and column[
                            1] != 'модуль_2']

        # Початкові координати для запису даних
        row = 3
        column = 4  # Стартова колонка (D)

        # Ширина стовпця
        column_width = 4.5

        for item in column_names:
            # Записуємо елемент у відповідну комірку
            cell = sheet.cell(row=row, column=column, value=item)

            # Встановлюємо вертикальне центрування, горизонтальне центрування та ширину стовпця
            alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
            cell.alignment = alignment
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width

            # Переходимо до наступної комірки
            column += 1

        # Після останнього елементу додаємо "Змістовий модуль 1"
        cell = sheet.cell(row=row, column=column, value="Змістовий модуль 1")
        alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
        cell.alignment = alignment
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width

        cursor.execute(f"PRAGMA table_info({pred}_2)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if
                        column[1] != 'Студенти' and column[1] != 'Н' and column[1] != 'модуль_1' and column[
                            1] != 'модуль_2']

        # Початкові координати для запису даних
        row = 3
        column = column + 1

        for item in column_names:
            # Записуємо елемент у відповідну комірку
            cell = sheet.cell(row=row, column=column, value=item)

            # Встановлюємо вертикальне центрування, горизонтальне центрування та ширину стовпця
            alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
            cell.alignment = alignment
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width

            # Переходимо до наступної комірки
            column += 1

        # Після останнього елементу додаємо "Змістовий модуль 1"
        cell = sheet.cell(row=row, column=column, value="Змістовий модуль 2")
        alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
        cell.alignment = alignment
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width

        column += 1
        # Записуємо "ІНДЗ"
        cell = sheet.cell(row=row, column=column, value="ІНДЗ")
        alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
        cell.alignment = alignment
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width

        # Переходимо до наступного стовпця
        column += 1

        # Записуємо "Підсумковий контроль"
        cell = sheet.cell(row=row, column=column, value='Підсумковий контроль')
        alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
        cell.alignment = alignment
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width

        # Переходимо до наступного стовпця
        column += 1

        # Записуємо "Кількість балів"
        cell = sheet.cell(row=row, column=column, value='Кількість балів')
        alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
        cell.alignment = alignment
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width

        # Переходимо до наступного стовпця
        column += 1

        # Записуємо "Оцінка ECTS"
        cell = sheet.cell(row=row, column=column, value='Оцінка ECTS')
        alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
        cell.alignment = alignment
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width


        cursor.execute(f"PRAGMA table_info({pred}_1)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns if column[1] != 'Студенти' and column[1] != 'Н']
        columns_str = ', '.join(column_names)
        elements = columns_str.split(", ")
        # Переміщуємо перший елемент (модуль_1) в кінець списку
        elements.append(elements.pop(0))
        columns_str = ", ".join([f"[{element}]" for element in elements])

        cursor.execute(f"PRAGMA table_info({pred}_2)")
        col = cursor.fetchall()
        col = [column[1] for column in col if column[1] != 'Студенти' and column[1] != 'Н']
        colum = ', '.join(col)
        ele = colum.split(", ")
        ele.append(ele.pop(0))
        colum = ", ".join([f"[{el}]" for el in ele])

        c = columns_str.split(", ")
        new_text = ", ".join(c[:-1])

        cursor.execute(f"SELECT {new_text} FROM {pred}_1 WHERE Студенти = ?", ("Дата",))
        data_1 = cursor.fetchall()
        data_1 = [item for tpl in data_1 for item in tpl]

        c_2 = colum.split(", ")
        new_text_2 = ", ".join(c_2[:-1])
        cursor.execute(f"SELECT {new_text_2} FROM {pred}_2 WHERE Студенти = ?", ("Дата",))
        data_2 = cursor.fetchall()
        data_2 = [item for tpl in data_2 for item in tpl]

        data = data_1 + data_2

        column = 4
        rowes = 4
        for date in data:
            # Записуємо елемент у відповідну комірку
            cell = sheet.cell(row=rowes, column=column, value=date)
            # Встановлюємо вертикальне центрування, горизонтальне центрування та ширину стовпця
            alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', textRotation=90)
            cell.alignment = alignment
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = column_width
            # Переходимо до наступної комірки
            column += 1

        row = 5

        for i in results:
            cursor.execute(f"SELECT {columns_str} FROM {pred}_1 WHERE Студенти = ?", (i,))
            res_1 = cursor.fetchall()
            res_1 = [item for tpl in res_1 for item in tpl]
            res_1 = [0 if item is None else item for item in res_1]

            cursor.execute(f"SELECT {colum} FROM {pred}_2 WHERE Студенти = ?", (i,))
            res_2 = cursor.fetchall()
            res_2 = [item for tpl in res_2 for item in tpl]
            res_2 = [0 if item is None else item for item in res_2]

            cursor.execute(f"SELECT Індз, {exam}, Загальна_кількість_балів  FROM {pred}_3 WHERE Студенти = ?", (i,))
            res_3 = cursor.fetchall()
            res_3 = [item for tpl in res_3 for item in tpl]

            res = res_1 + res_2 + res_3

            column = 4
            for item in res:
                # Записуємо елемент у відповідну комірку
                cell = sheet.cell(row=row, column=column, value=item)
                # Встановлюємо вертикальне центрування, горизонтальне центрування та ширину стовпця
                alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')

                # Переходимо до наступної комірки
                column += 1

            row += 1

        # Зберігаємо зміни
        workbook.save(f'{pred1}.xlsx')
        workbook = openpyxl.load_workbook(f'{pred1}.xlsx')
        with open(f'{pred1}.xlsx', 'rb') as file:
            bot.send_document(message.chat.id, file)



        workbook.close()

        # Видаляємо файл після збереження
        os.remove(f'{pred1}.xlsx')

class Create_jurnal:
    def __init__(self, bot):
        self.bot = bot

    @staticmethod
    def jurnal2_1(message, user_grypa):

        # Створення бази даних з назвою групи
        conn = sqlite3.connect(f"{user_grypa}.db")
        cursor = conn.cursor()

        # Створення таблиці STUDENTY
        cursor.execute("CREATE TABLE IF NOT EXISTS STUDENTY (Студенти TEXT)")

        # Створення таблиці Предмети
        cursor.execute("CREATE TABLE IF NOT EXISTS Предмети (Предмети TEXT,Кредити_ЄКТС TEXT,Форма_підсумкового_контролю TEXT,Закритий_модуль_1 TEXT,Закритий_модуль_2 TEXT, Закритий_предмет TEXT)")
        #Створення лекційного тижня
        cursor.execute("CREATE TABLE лекційний_Понеділок (Студенти TEXT,Назвапредмету1 TEXT,Назвапредмету2 TEXT,Назвапредмету3 TEXT,Назвапредмету4 TEXT)")
        cursor.execute("CREATE TABLE лекційний_Вівторок (Студенти TEXT,Назвапредмету1 TEXT,Назвапредмету2 TEXT,Назвапредмету3 TEXT,Назвапредмету4 TEXT)")
        cursor.execute("CREATE TABLE лекційний_Середа (Студенти TEXT,Назвапредмету1 TEXT,Назвапредмету2 TEXT,Назвапредмету3 TEXT,Назвапредмету4 TEXT)")
        cursor.execute("CREATE TABLE лекційний_Четвер (Студенти TEXT,Назвапредмету1 TEXT,Назвапредмету2 TEXT,Назвапредмету3 TEXT,Назвапредмету4 TEXT)")
        cursor.execute("CREATE TABLE лекційний_Пятниця (Студенти TEXT,Назвапредмету1 TEXT,Назвапредмету2 TEXT,Назвапредмету3 TEXT,Назвапредмету4 TEXT)")
        cursor.execute("CREATE TABLE лекційний_Субота (Студенти TEXT,Назвапредмету1 TEXT,Назвапредмету2 TEXT,Назвапредмету3 TEXT,Назвапредмету4 TEXT)")
        cursor.execute("CREATE TABLE Лекційний_Тиждень (Лекційний_Тиждень TEXT)")
        cursor.execute(f'INSERT INTO Лекційний_Тиждень VALUES (NULL)')
        # Збереження змін до бази даних
        conn.commit()

        # Закриття підключення до бази даних
        conn.close()
        bot.send_message(message.chat.id, "Будь ласка, надішліть мені список студентів вашої групи за таким зразком")
        bot.send_message(message.chat.id, "Прізвище ім'я по-батькові\nПрізвище ім'я по-батькові\nПрізвище ім'я по-батькові\nПрізвище ім'я по-батькові\nПрізвище ім'я по-батькові")
        bot.register_next_step_handler(message, Create_jurnal.jurnal2_2, user_grypa)

    @staticmethod
    def jurnal2_2(message, user_grypa):
        text = message.text


        # Знаки, які ми хочемо видалити
        characters_to_remove = '().:;\|/,+#*%@$&?!~"1234567890'
        text = ''.join(filter(lambda x: not x.isdigit(), text))
        # Видаляємо знаки з тексту
        for char in characters_to_remove:
            text = text.replace(char, '')
        #
        abzac = text.split('\n')
        num_paragraphs = len(abzac)
        num_paragraphs = num_paragraphs - 1
        text = text.split('\n')
        text.sort()
        pravelni = []
        nepravelni = []
        for i in range(len(text)):
            reversed_string = text[i].rsplit()
            if len(reversed_string) < 2 or len(reversed_string) > 3:
                p = ' '.join(reversed_string)
                nepravelni.append(p)
            elif len(reversed_string) == 2 or len(reversed_string) == 3:
                p = ' '.join(reversed_string)
                pravelni.append(p)
        if len(nepravelni) > 0:
            bot.send_message(message.chat.id, "Ви ввели щось не правильно, попробуйте ще раз, мабуть ви вказали ПІБ студентів через пробіл, а потрібно через абзац")
            bot.register_next_step_handler(message, Create_jurnal.jurnal2_2, user_grypa)
        else:
            conn = sqlite3.connect("users.db")
            cursor = conn.cursor()
            cursor.execute("SELECT ПІП FROM Email_Base")
            rows = cursor.fetchall()
            pip = [row[0] for row in rows]
            mana = []
            mananot = []
            for student in text:
                if student in pip:
                    mana.append(student)
                elif student not in pip:
                    mananot.append(student)
            if len(mananot) > 0:
                str1 = ''
                for i in mananot:
                    str1 += i + "\n"
                bot.send_message(message.chat.id,f"Ви ввели не правильно ось ці ПІБ\n\n\n{str1}\n\nвведіть ще раз весь список групи через абзац кожен студент")
                bot.register_next_step_handler(message, Create_jurnal.jurnal2_2, user_grypa)
            elif len(mananot) == 0:
                Create_jurnal.jurnal2_3(message, user_grypa, text)

    @staticmethod
    def jurnal2_3(message,user_grypa,text):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Так все вірно')
        item2 = types.KeyboardButton('Редагувати список')
        markup.add(item1)
        markup.add(item2)
        bot.send_message(message.chat.id, "Будь ласка, перевірте чи ви правильно надіслали мені список своїх одногрупників".format(message.from_user), reply_markup=markup)
        bot.register_next_step_handler(message,Create_jurnal.jurnal2_4,user_grypa,text)

    @staticmethod
    def jurnal2_4(message,user_grypa,text):
        vid = message.text


        if vid == 'Так все вірно':

            bot.send_message(message.chat.id, "Вношу в базу...", reply_markup=telebot.types.ReplyKeyboardRemove())
            time.sleep(2)
            Create_jurnal.jurnal2_5(message, user_grypa, text)

        elif vid == 'Редагувати список':
            bot.send_message(message.chat.id, "Надішліть ще раз список")
            bot.register_next_step_handler(message, Create_jurnal.jurnal2_2, user_grypa)

        else:
            bot.send_message(message.chat.id, "Такого варіанту відповіді немає виберіть із наявих кнопок")
            bot.register_next_step_handler(message, Create_jurnal.jurnal2_4, user_grypa, text)

    @staticmethod
    def jurnal2_5(message,user_grypa,text):
        conn = sqlite3.connect(f"{user_grypa}.db")
        cursor = conn.cursor()


        for row in text:
            cursor.execute("INSERT INTO STUDENTY (Студенти) VALUES (?)", (row,))

        conn.commit()
        bot.send_message(message.chat.id, "Збережено✅")
        bot.send_message(message.chat.id, 'А тепер перейдемо до створення журналу предметів, надішліть мені повні назви своїх навчальних дисциплін цього семестру за такою формою:\n"Назва предмету" - (кількість кредитів ЄКТС), (форма підсумкового контролю)\n"Назва предмету" - (кількість кредитів ЄКТС), (форма підсумкового контролю)\n‼️УВАГА‼️️:\nБудь уважні та робіть пробіли між словами та символами як у зразку⬇️\n\nЗРАЗОК :\nМатематичний аналіз - 4, залік\n"Назва предмету" - (кількість кредитів ЄКТС),(форма підсумкового контролю)\n........\nПрограмування - 5, екзамен\n\nНадсилати ПІБ викладачів, які ведуть у вас цей предмет, не потрібно🫠')
        bot.register_next_step_handler(message, Create_jurnal.jurnal2_6, user_grypa)

    @staticmethod
    def jurnal2_6(message,user_grypa):
        pred = message.text
        predmety = pred
        lines = pred.split('\n')
        list = []
        subject_list = []

        for i in lines:
            elements = i.split(',')
            if len(elements) != 2:

                list.append("Помилка: Неправильний формат рядка -" + i)
                continue
            a = elements[0].split("-")
            if len(a) != 2:

                list.append("Помилка: Неправильний формат рядка (не вдалося розбити по дефісу) попробуйте предмети в яких є дефіс написати окремо  'Game дизайн', 'Smart технології'\n‼️‼️\nЯкщо виникли труднощі напишіть в /support,  щоб вам додали предмет, який не можете добавити, а його пропустіть і не додавайте до основного списку\n‼️‼️" + i)
                continue
            t = tuple(k.strip() for k in a)
            subject_list.append(t + (elements[1].strip(),))

        list = tuple(list)
        message_text = "\n\n".join(list)



        message_text2 = "\n\n".join([f'{item[0]} - {item[1]}, {item[2]}' for item in subject_list])

        if len(list) > 0:
            if len(subject_list) > 0:
                bot.send_message(message.chat.id,f"{message_text}\n тут виведені предмети, які Ви ввели не правильно\n\n{message_text2}\nА ось тут правильні\nВведіть будь ласка ще раз свої предмети за зразком")
                bot.register_next_step_handler(message, Create_jurnal.jurnal2_6, user_grypa)
            elif len(subject_list) == 0:
                bot.send_message(message.chat.id, f"{message_text}\n тут виведені предмети, які Ви ввели не правильно.Введіть будь ласка ще раз свої предмети за зразком")
                bot.register_next_step_handler(message, Create_jurnal.jurnal2_6, user_grypa)
        else:
            Create_jurnal.jurnal2_6_1(message, user_grypa, predmety)

    @staticmethod
    def jurnal2_6_1(message,user_grypa,predmety):
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton('Так все вірно')
        item2 = types.KeyboardButton('Редагувати')
        markup.add(item1)
        markup.add(item2)
        bot.send_message(message.chat.id,"Перевірте чи ви правильно надіслали усі навчальні дисципліни та їх форму підсумкового контролю і кредити ",
                         reply_markup=markup)
        bot.send_message(message.chat.id, f"<code>{predmety}</code>", parse_mode=ParseMode.HTML)
        bot.register_next_step_handler(message, Create_jurnal.jurnal2_7, user_grypa, predmety)

    @staticmethod
    def jurnal2_7(message,user_grypa, predmety):
        vid = message.text

        if vid == 'Так все вірно':

            bot.send_message(message.chat.id, "Провіряю ваші предмети...", reply_markup=telebot.types.ReplyKeyboardRemove())
            time.sleep(2)
            Create_jurnal.jurnal2_8(message, user_grypa, predmety)


        elif vid == 'Редагувати':
            bot.send_message(message.chat.id,"Будь ласка надішліть мені ще раз ваші навчальні дисципліни і переконайтеся чи все вірно)", reply_markup=telebot.types.ReplyKeyboardRemove())
            bot.register_next_step_handler(message,Create_jurnal.jurnal2_6, user_grypa)
        else:
            bot.send_message(message.chat.id,"Такого варіанту відповіді немає виберіть із наявих кнопок")
            bot.register_next_step_handler(message,Create_jurnal.jurnal2_7,user_grypa, predmety)

    @staticmethod
    def jurnal2_8(message, user_grypa, predmety):
        subjects = predmety.split('\n')
        subject_list = []


        for subject in subjects:
            subject_parts = subject.split(' - ')
            if len(subject_parts) == 2:
                subject_name = subject_parts[0].strip().lower()
                credits_exam_parts = subject_parts[1].split(',')

                if len(credits_exam_parts) == 2:
                    credits = credits_exam_parts[0].strip()
                    exam_type = credits_exam_parts[1].strip().lower()

                    subject_tuple = (subject_name, credits, exam_type)
                    subject_list.append(subject_tuple)
        q=0
        for i in subject_list:
            if i[2] != "залік" and i[2] != 'екзамен':
                q += 1
                bot.send_message(message.chat.id, "Бачу в предметі " + i[0] + " Ви ввели не вірну Форму підсумкового контролю будь ласка вкажіть ще раз всі предмети")

        if q > 0:
            bot.register_next_step_handler(message, Create_jurnal.jurnal2_6, user_grypa)

        else:
            str1 = ''
            conn = sqlite3.connect("users.db")
            cursor = conn.cursor()
            cursor.execute("SELECT ПредметиФізмату FROM ПредметиФізмату")
            result = cursor.fetchall()
            result = [item.lower() for tpl in result for item in tpl]

            for i in subject_list:
                if i[0] not in result:
                    str1 += i[0] + ' - ' + 'Такого предмету не існує' + '\n'
            if len(str1) > 0:
                bot.send_message(message.chat.id, str1)
                bot.send_message(message.chat.id, "Введіть всі свої предмети ще раз")
                bot.register_next_step_handler(message, Create_jurnal.jurnal2_6, user_grypa)
            else:
                conn = sqlite3.connect(f"{user_grypa}.db")
                cursor = conn.cursor()

                for subject_tuple in subject_list:
                    subject_name, credits, exam_type = subject_tuple

                    cursor.execute("INSERT INTO Предмети (Предмети, Кредити_ЄКТС, Форма_підсумкового_контролю) VALUES (?, ?, ?)", (subject_name, credits, exam_type))

                conn.commit()
                conn.close()
                bot.send_message(message.chat.id, "Створюю таблиці з вашими предметами...")
                Create_jurnal.jurnal2_9(message, user_grypa)

    @staticmethod
    def jurnal2_9(message,user_grypa):
        conn = sqlite3.connect(f"{user_grypa}.db")
        cursor = conn.cursor()
        cursor.execute("SELECT Предмети, Форма_підсумкового_контролю FROM Предмети")
        result = cursor.fetchall()
        result = [(item[0].replace(" ", "_"), item[1]) for item in result]

        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()

        for row in result:
            pre, exam_type = row
            cursor.execute(f'CREATE TABLE "{pre}_1" ( Студенти TEXT, модуль_1 TEXT, Н TEXT, [тема 0] TEXT,[тема 1] TEXT, [тема 2] TEXT, [тема 3] TEXT, [тема 4] TEXT, [тема 5] TEXT, [тема 6] TEXT, [тема 7] TEXT,  [тема 8] TEXT, [тема 9] TEXT)')
            cursor.execute(f'INSERT INTO "{pre}_1" ("Студенти", модуль_1, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Дата", NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')
            cursor.execute(f'INSERT INTO "{pre}_1" ("Студенти", модуль_1, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Тип заняття", NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')

            cursor.execute(f'CREATE TABLE "{pre}_2" ( Студенти TEXT, модуль_2 TEXT, Н TEXT, [тема 0] TEXT,[тема 1] TEXT, [тема 2] TEXT, [тема 3] TEXT, [тема 4] TEXT, [тема 5] TEXT, [тема 6] TEXT, [тема 7] TEXT,  [тема 8] TEXT, [тема 9] TEXT)')
            cursor.execute(f'INSERT INTO "{pre}_2" ("Студенти", модуль_2, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Дата",NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')
            cursor.execute(f'INSERT INTO "{pre}_2" ("Студенти", модуль_2, Н, [тема 0], [тема 1], [тема 2], [тема 3], [тема 4], [тема 5], [тема 6], [тема 7], [тема 8], [тема 9]) VALUES ("Тип заняття", NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)')

            cursor.execute(f'CREATE TABLE "{pre}_3" ("Студенти" TEXT,Індз TEXT, модуль_1 TEXT, модуль_2 TEXT, [{exam_type}] TEXT,Індивідуальні_години TEXT, Загальна_кількість_балів TEXT)')


            cursor.execute(f'CREATE TABLE "{pre}_Індивідуальні_години" (Студенти TEXT)')



            cursor.execute(f'CREATE TABLE "{pre}_Студенти" ("Студенти" TEXT,Індз TEXT, модуль_1 TEXT, модуль_2 TEXT, [{exam_type}] TEXT,Індивідуальні_години TEXT, Загальна_кількість_балів TEXT)')
            #cursor.execute(f'INSERT INTO "{pre}_Студенти" ("Студенти", [модуль 1], [модуль 2], [індз], [підсумковий контроль], [загальна кількість балів]) VALUES ("Дата", NULL, NULL, NULL, NULL, NULL)')

            for student in students:
                cursor.execute(f"INSERT INTO \"{pre}_1\" (\"Студенти\") VALUES (?)", (str(student[0]),))
                cursor.execute(f"INSERT INTO \"{pre}_2\" (\"Студенти\") VALUES (?)", (str(student[0]),))
                cursor.execute(f"INSERT INTO \"{pre}_3\" (\"Студенти\") VALUES (?)", (str(student[0]),))
                cursor.execute(f"INSERT INTO \"{pre}_Студенти\" (\"Студенти\") VALUES (?)", (str(student[0]),))
                cursor.execute(f"INSERT INTO \"{pre}_Індивідуальні_години\" (\"Студенти\") VALUES (?)", (str(student[0]),))
            days = ["Понеділок","Вівторок","Середа","Четвер","Пятниця","Субота"]
            for student in students:
                for i in days:
                    cursor.execute(f"INSERT INTO \"лекційний_{i}\" (\"Студенти\") VALUES (?)", (str(student[0]),))
        conn.commit()
        conn.close()
        bot.send_message(message.chat.id, "Таблиці успішно створенні✅")
        message_handler_start(message)



def redaguvanna(message,user_id):
    den = message.text
    if den == "🔙Назад":
        menu_starostam(message)
    elif den not in ['Понеділок','Вівторок','Середа','Четвер','П\'ятниця']:
        bot.send_message(user_id, "Ви ввели не правильний день, будь ласка виберіть день із наявних кнопок: ")
        bot.register_next_step_handler(message, redaguvanna, user_id)
    else:
        key = types.ReplyKeyboardMarkup(resize_keyboard=True)
        conn = sqlite3.connect('users.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT grypa FROM login_id WHERE id = {user_id}")
        user_grypa = cursor.fetchone()[0]
        back = types.KeyboardButton('🔙Назад')
        key.add(back)
        bot.send_message(message.chat.id, f"Ви вибрали {den}, будь ласка надішліть відредагований розкла.\nЗА ТАКИМ ЗРАЗКОМ!!!", reply_markup=key)
        #Продовження допиши як правильно вставляти
        bot.send_message(message.chat.id,'Напишіть назви пар відповідно до їх порядку, якщо у вас вікно тобто не має пари, то можете пропусти її нумерацію, або поставити прочерк')
        redaguvanna2(message, user_id, user_grypa, den)
def redaguvanna2(message, user_id,user_grypa,den):
    days = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'П\'ятниця']
    den_123 = days.index(den)

    # Встановлюємо з'єднання з базою даних
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()

    cursor.execute(f'SELECT * FROM rosklad_{user_grypa}')
    # Отримуємо результат запиту (перший рядок таблиці)
    first_row = cursor.fetchall()[den_123]

    mess = ''
    for j, item in enumerate(first_row):
        if item is not None:
            mess += f"{j + 1}. {item}\n"

    bot.send_message(user_id, f'\n<code>{mess}</code>', parse_mode=ParseMode.HTML)
    bot.register_next_step_handler(message, redaguvanna3, user_grypa, user_id, den_123)
    # Закриваємо з'єднання з базою даних
    cursor.close()
    conn.close()
def redaguvanna3(message, user_grypa, user_id, den_123):
    text = message.text  # Отримуємо текст повідомлення
    if text == '🔙Назад':
        menu_starostam(message)
    else:
        try:


            # Розбиваємо текст на окремі рядки
            lines = text.split('\n')

            result = [None] * 5  # Результат (кортеж) з початковим значенням None

            for line in lines:
                # Розбиваємо рядок на індекс і значення
                index, value = line.split('. ', 1)

                # Перевіряємо, чи індекс є числом від 1 до 5
                if index.isdigit() and 1 <= int(index) <= 5:
                    result[int(index) - 1] = value

            # Виводимо результат (кортеж)
            result = tuple(result)
            bot.send_message(user_id, "Розклад успішно оновленно")
            redaguvanna4(message, user_grypa, user_id, den_123, result )

        except Exception as e:
           bot.send_message(user_id, 'Ви ввели не по зразку, будь ласка надішліть мені ще раз відредагований розклад за зразком')
           bot.send_message(user_id,'<code>1. (Назва пари)\n2. (Назва пари)\n3. (Назва пари)\n4. (Назва пари)\n5. (Назва пари)</code>', parse_mode=ParseMode.HTML)
           bot.register_next_step_handler(message, redaguvanna3, user_grypa, user_id, den_123)
def redaguvanna4(message, user_grypa, user_id, den_123, result):
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute(f'SELECT * FROM rosklad_{user_grypa}')
    rows = cursor.fetchall()
    rows[den_123] = result
    cursor.execute(f'DELETE FROM rosklad_{user_grypa}')
    cursor.executemany(f'INSERT INTO rosklad_{user_grypa} VALUES (?,?,?,?,?)', rows)
    conn.commit()
    cursor.close()
    conn.close()
    message_handler_start(message)
@bot.callback_query_handler(func=lambda call: True)
def handle_callback_query(call):
    # Отримання інформації про викладача з бази даних
    conn = sqlite3.connect('users.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM Teachers WHERE Викладач=?", (call.data,))
    teacher_info = cursor.fetchone()
    conn.close()

    # Формування повідомлення про викладача
    teacher_name = teacher_info[0]
    phone_number = teacher_info[1]
    email = teacher_info[2]
    full_name = teacher_info[3]
    # Відправка повідомлення про викладача
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text = f"Викладач: {full_name}\nТелефон: <code>{phone_number}</code>\nПошта: {email}",parse_mode = ParseMode.HTML)



class Rozklad:
    def __init__(self, bot):
        self.bot = bot

    @staticmethod
    def rozklad_par_0_1(message, user_id,user_grypa):
        connect = sqlite3.connect('users.db')
        cursor = connect.cursor()
        cursor.execute(f"SELECT * FROM rosklad_{user_grypa}")
        dates = cursor.fetchall()
        days = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'П`ятниця']
        times = ['8:00-9:20', '9:35-10:55', '11:10-12:30', '12:45-14:05', '14:20-15:40']
        mess = ''
        for i, date in enumerate(dates):
            mess += days[i] + ':\n'
            for j, item in enumerate(date):
                if item is not None:
                    mess += f"{times[j]}: {item}\n"
            mess += '\n'


        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('Переглянути інші'))
        markup.add(types.KeyboardButton('🔙Назад'))
        bot.send_message(user_id, "Розклад пар для вашої групи " + user_grypa.replace('_', '-') + ':\n\n' + mess, reply_markup=markup)
        bot.register_next_step_handler(message, Rozklad.rozklad_par_0_2, user_id)

    @staticmethod
    def rozklad_par_0_2(message, user_id):
        text = message.text
        if text == 'Переглянути інші':
            Rozklad.rozklad_par_0(message, user_id)
        elif text == '🔙Назад':
            message_handler_start(message)
        else:
            bot.send_message(user_id, "Такого варіанту відповіді немає, оберіть ще раз")
            bot.register_next_step_handler(message, Rozklad.rozklad_par_0_2, user_id)

    @staticmethod
    def rozklad_par_0(message,user_id):
        connect = sqlite3.connect('users.db')
        cursor = connect.cursor()
        cursor.execute("SELECT Групи FROM Групи")
        rows = cursor.fetchall()
        gryps = [row[0] for row in rows]
        columns = 3
        gryps_per_column = (len(gryps) + columns - 1) // columns
        gryps_divided = [gryps[i:i + gryps_per_column] for i in range(0, len(gryps), gryps_per_column)]

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.add(types.KeyboardButton('🔙Назад'))
        # Додавання груп до розмітки по 3 в кожному рядку
        for gryp_column in gryps_divided:
            markup.add(*gryp_column)

        connect.close()
        bot.send_message(message.chat.id, 'Виберіть групу:', reply_markup=markup)
        bot.register_next_step_handler(message, Rozklad.rozklad_par,user_id,gryps)

    @staticmethod
    def rozklad_par(message,user_id,gryps):
        data = message.text
        if message.text == '🔙Назад':
            message_handler_start(message)

        elif data not in gryps:
            bot.send_message(message.chat.id, "Ви ввели не правильну групу, будь ласка виберіть групу із наявних кнопок: ")
            bot.register_next_step_handler(message, Rozklad.rozklad_par,user_id,gryps)
        else:
            data = message.text.upper().replace('-', '_')
            Rozklad.rozklad_par2(message, data, user_id,gryps)

    @staticmethod
    def rozklad_par2(message,data,user_id,gryps):

        connect = sqlite3.connect('users.db')
        cursor = connect.cursor()
        cursor.execute(f"SELECT * FROM rosklad_{data}")
        dates = cursor.fetchall()
        days = ['Понеділок', 'Вівторок', 'Середа', 'Четвер', 'П`ятниця']
        times = ['8:00-9:20', '9:35-10:55', '11:10-12:30', '12:45-14:05', '14:20-15:40']
        mess = ''
        for i, date in enumerate(dates):
            mess += days[i] + ':\n'
            for j, item in enumerate(date):
                if item is not None:
                    mess += f"{times[j]}: {item}\n"
            mess += '\n'

        bot.send_message(user_id,"Розклад пар для групи " + data.replace('_', '-') + ':\n\n' + mess)
        bot.register_next_step_handler(message, Rozklad.rozklad_par,user_id ,gryps)


def jurnal_prerglad_ocinok_dla_starost(message,db_filename,user_grypa,subject, table):
    conn = sqlite3.connect(db_filename)
    cursor = conn.cursor()

    # Формуємо назву таблиці у форматі {subject}_{table}
    table_name = f'{subject}_{table}'

    # Отримуємо назви стовпців таблиці
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = cursor.fetchall()
    column_names = [column[1] for column in columns if
                    column[1] != 'Студенти' and column[1] != 'Н']

    # Створюємо клавіатуру з кнопками зі списку column_names
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton('🔙Назад'))
    for column_name in column_names:
        markup.add(column_name)

    bot.send_message(message.chat.id,f"Оберіть тему, в модулі {table} з якої ви хочете переглянути оцінки своїх студентів або одразу із цілого модуля:",reply_markup=markup)
    bot.register_next_step_handler(message, jurnal_prerglad_ocinok_dla_starost_2, db_filename,user_grypa,subject, table, column_names)
def jurnal_prerglad_ocinok_dla_starost_2(message,db_filename,user_grypa,subject, table, column_names):
    text = message.text
    if text == '🔙Назад':
        message_handler_start(message)
    elif text not in column_names:
        bot.send_message(message.chat.id, "Ви вибрали не вірну тему оберіть ще раз")
        bot.register_next_step_handler(message, jurnal_prerglad_ocinok_dla_starost_2, db_filename, user_grypa, subject, table,column_names)

    else:
        table_name = f'{subject}_{table}'
        conn = sqlite3.connect(db_filename)
        cursor = conn.cursor()
        cursor.execute("SELECT Студенти FROM STUDENTY")
        students = cursor.fetchall()
        # Формуємо повідомлення зі списком студентів
        students_list = "\n".join([student[0] + ' - 0' for student in students])

        cursor.execute(f"SELECT [{text}] FROM {table_name} ")
        ocinky = cursor.fetchall()
        ocinky = ocinky[2:]
        ocinky = [item for tpl in ocinky for item in tpl]
        ocinky2 = []
        for i in ocinky:
            if i == None:
                ocinky2.append("0")
            else:
                ocinky2.append(i)

        results = []
        students_list1 = students_list.split("\n")
        for k, i in enumerate(students_list1):
            results.append(i + ocinky2[k])

        gem = "\n".join([row for row in results])

        bot.send_message(message.chat.id, f"Ось оцінки із предмету {subject.replace('_', ' ')} із модуля {table}\n{gem}")
        bot.register_next_step_handler(message, jurnal_prerglad_ocinok_dla_starost, db_filename, user_grypa, subject,table)


bot.polling(none_stop=True)
